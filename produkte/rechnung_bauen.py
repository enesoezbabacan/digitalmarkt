"""
Erzeugt die Rechnungsvorlage mit allen Pflichtangaben nach § 14 UStG.

Verkaufsprodukt für den Marktplatz. Zielgruppe: Kleinunternehmer und
Solo-Selbstständige, die Rechnungen selbst schreiben.

Kern des Produkts: Eine fehlende Pflichtangabe macht die Rechnung für den
Kunden nicht vorsteuerabzugsfähig — er verlangt dann eine Korrektur, im
schlimmsten Fall zahlt er erst danach. Die Vorlage prüft die Pflichtangaben
automatisch und meldet, was fehlt.

Aufbau:
  - Anleitung         Bedienung, Pflichtangaben, Haftungshinweis
  - Meine Daten       einmalig ausfüllen, erscheint auf jeder Rechnung
  - Rechnung          das Formular zum Ausfüllen und Drucken
  - Pflichtangaben    Prüfliste, die automatisch anschlägt
  - Rechnungsbuch     fortlaufende Nummern und Zahlungsüberwachung
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCHRIFT = "Arial"
DUNKEL = "1F3864"
HELL = "D9E2F3"
GELB = "FFF2CC"
GRAU = "F2F2F2"
GRUEN = "E2EFDA"
ROT = "FCE4E4"

EURO = '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'
DATUM = "TT.MM.JJJJ"

rand = Side(style="thin", color="BFBFBF")
RAHMEN = Border(left=rand, right=rand, top=rand, bottom=rand)

POSITIONEN = 15   # Positionszeilen auf der Rechnung
BUCH_ZEILEN = 200  # Zeilen im Rechnungsbuch


def titel(ws, zelle, text, groesse=14, farbe=DUNKEL):
    ws[zelle] = text
    ws[zelle].font = Font(name=SCHRIFT, size=groesse, bold=True, color=farbe)


def eingabe(zelle):
    zelle.font = Font(name=SCHRIFT, size=11, color="0000FF")
    zelle.fill = PatternFill("solid", fgColor=GELB)
    zelle.border = RAHMEN


wb = Workbook()

# ---------------------------------------------------------------------------
# Anleitung
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

titel(ws, "B2", "Rechnungsvorlage mit allen Pflichtangaben", 18)

text = [
    ("", None),
    ("So gehst du vor", "kopf"),
    ("1. Blatt „Meine Daten“ einmalig ausfüllen — das erscheint auf jeder Rechnung.", None),
    ("2. Im Blatt „Rechnung“ Kunde, Nummer, Datum und Positionen eintragen.", None),
    ("3. Blatt „Pflichtangaben“ prüfen — dort steht, ob etwas fehlt.", None),
    ("4. Rechnung als PDF speichern: Datei → Drucken → als PDF sichern.", None),
    ("5. Im „Rechnungsbuch“ eintragen, damit die Nummern lückenlos bleiben.", None),
    ("", None),
    ("Warum die Pflichtangaben wichtig sind", "kopf"),
    ("Fehlt auch nur eine Pflichtangabe nach § 14 UStG, darf dein Kunde die", None),
    ("Rechnung nicht als Vorsteuer geltend machen. Er wird sie zurückschicken", None),
    ("und eine Korrektur verlangen — im schlechtesten Fall zahlt er erst danach.", None),
    ("Das Blatt „Pflichtangaben“ prüft das automatisch für dich.", None),
    ("", None),
    ("Die Pflichtangaben nach § 14 Abs. 4 UStG", "kopf"),
    ("   1. Vollständiger Name und Anschrift des leistenden Unternehmers", None),
    ("   2. Vollständiger Name und Anschrift des Leistungsempfängers", None),
    ("   3. Steuernummer oder USt-IdNr. des Leistenden", None),
    ("   4. Ausstellungsdatum", None),
    ("   5. Fortlaufende, einmalige Rechnungsnummer", None),
    ("   6. Menge und Art der gelieferten Gegenstände oder Leistungen", None),
    ("   7. Zeitpunkt der Lieferung oder Leistung", None),
    ("   8. Entgelt, aufgeschlüsselt nach Steuersätzen", None),
    ("   9. Steuersatz und Steuerbetrag — oder Hinweis auf Steuerbefreiung", None),
    ("", None),
    ("Besonderheit für Kleinunternehmer", "kopf"),
    ("Als Kleinunternehmer nach § 19 UStG weist du KEINE Umsatzsteuer aus.", None),
    ("Stattdessen ist ein Hinweis auf die Steuerbefreiung Pflicht. Die Vorlage", None),
    ("setzt ihn automatisch, sobald du im Blatt „Meine Daten“ auf JA stellst.", None),
    ("", None),
    ("Fortlaufende Nummern", "kopf"),
    ("Rechnungsnummern müssen fortlaufend und einmalig sein — keine Lücken,", None),
    ("keine Doppelvergabe. Empfohlenes Muster: 2026-001, 2026-002, 2026-003 …", None),
    ("Das Rechnungsbuch warnt dich, wenn eine Nummer doppelt vorkommt.", None),
    ("", None),
    ("Aufbewahrung", "kopf"),
    ("Rechnungen musst du 10 Jahre aufbewahren, unveränderbar. Ein PDF auf", None),
    ("der Festplatte reicht, solange du es nicht nachträglich änderst.", None),
    ("Die Excel-Datei selbst ist KEIN gültiger Beleg — nur das erzeugte PDF.", None),
    ("", None),
    ("Wichtiger Hinweis", "warnung"),
    ("Diese Vorlage ist eine Arbeitshilfe, KEINE Steuerberatung und keine", None),
    ("Rechtsberatung. Für die Richtigkeit deiner Rechnungen bist ausschließlich", None),
    ("du verantwortlich. Bei Auslandsgeschäften, Reverse Charge, Bauleistungen", None),
    ("oder Differenzbesteuerung gelten zusätzliche Regeln, die hier nicht", None),
    ("abgebildet sind. Im Zweifel: frag deinen Steuerberater.", None),
    ("", None),
    ("Rechtsstand: August 2026. Steuerrecht ändert sich — prüfe die Angaben.", None),
]

r = 3
for inhalt, art in text:
    z = ws.cell(row=r, column=2, value=inhalt)
    if art == "kopf":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color=DUNKEL)
    elif art == "warnung":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color="C00000")
    else:
        z.font = Font(name=SCHRIFT, size=11)
    r += 1

# ---------------------------------------------------------------------------
# Meine Daten
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Meine Daten")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 44

titel(ws, "B2", "Meine Daten", 16)
ws["B3"] = "Einmalig ausfüllen. Diese Angaben erscheinen automatisch auf jeder Rechnung."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

felder = [
    ("Name / Firma", "Max Mustermann"),
    ("Straße und Hausnummer", "Musterstraße 12"),
    ("PLZ und Ort", "78234 Engen"),
    ("Telefon", "07733 123456"),
    ("E-Mail", "info@beispiel.de"),
    ("Steuernummer", "12/345/67890"),
    ("USt-IdNr. (falls vorhanden)", ""),
    ("Kleinunternehmer § 19 UStG?", "JA"),
    ("Umsatzsteuersatz in % (falls nein)", 19),
    ("Bank / Institut", "Musterbank"),
    ("IBAN", "DE00 0000 0000 0000 0000 00"),
    ("BIC", "MUSTDEFFXXX"),
    ("Zahlungsziel in Tagen", 14),
]

r = 5
for beschriftung, beispiel in felder:
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11, bold=True)
    b.border = RAHMEN
    w = ws.cell(row=r, column=3, value=beispiel)
    eingabe(w)
    r += 1

ws["B19"] = "Bei „Kleinunternehmer“ genau JA oder NEIN eintragen — davon hängt ab,"
ws["B19"].font = Font(name=SCHRIFT, size=10, italic=True, color="C00000")
ws["B20"] = "ob auf der Rechnung Umsatzsteuer ausgewiesen wird."
ws["B20"].font = Font(name=SCHRIFT, size=10, italic=True, color="C00000")

# ---------------------------------------------------------------------------
# Rechnung
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Rechnung")
ws.sheet_view.showGridLines = False
for spalte, breite in zip("ABCDEF", [3, 34, 12, 10, 16, 18]):
    ws.column_dimensions[spalte].width = breite

# Absenderzeile (Pflichtangabe 1)
ws["B2"] = '=\'Meine Daten\'!C5&" · "&\'Meine Daten\'!C6&" · "&\'Meine Daten\'!C7'
ws["B2"].font = Font(name=SCHRIFT, size=8, color="595959")

# Empfänger (Pflichtangabe 2)
ws["B4"] = "Rechnungsempfänger"
ws["B4"].font = Font(name=SCHRIFT, size=9, bold=True, color="595959")
for i, platzhalter in enumerate(
    ["Kunde GmbH", "Kundenstraße 5", "10115 Berlin"], start=5
):
    z = ws.cell(row=i, column=2, value=platzhalter)
    eingabe(z)

titel(ws, "B10", "RECHNUNG", 22)

# Kopfdaten (Pflichtangaben 4, 5, 7)
kopf = [
    ("Rechnungsnummer", "2026-001"),
    ("Rechnungsdatum", "2026-08-05"),
    ("Leistungsdatum / -zeitraum", "Juli 2026"),
    ("Kundennummer (optional)", "K-100"),
]
r = 12
for beschriftung, beispiel in kopf:
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=10, bold=True)
    w = ws.cell(row=r, column=3, value=beispiel)
    eingabe(w)
    if beschriftung == "Rechnungsdatum":
        w.number_format = DATUM
    r += 1

# Positionen (Pflichtangabe 6)
kopfzeile_pos = 18
for i, (ueberschrift, breite) in enumerate(
    zip(["Beschreibung", "Menge", "Einheit", "Einzelpreis", "Gesamt"], [34, 12, 10, 16, 18]),
    start=2,
):
    z = ws.cell(row=kopfzeile_pos, column=i, value=ueberschrift)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.alignment = Alignment(horizontal="center")
    z.border = RAHMEN

erste = kopfzeile_pos + 1
letzte = kopfzeile_pos + POSITIONEN

# Beispielposition
ws.cell(row=erste, column=2, value="Beratungsleistung Projekt Muster")
ws.cell(row=erste, column=3, value=10)
ws.cell(row=erste, column=4, value="Std.")
ws.cell(row=erste, column=5, value=85.00)

for r in range(erste, letzte + 1):
    for c in range(2, 7):
        z = ws.cell(row=r, column=c)
        z.border = RAHMEN
        z.font = Font(name=SCHRIFT, size=10)
        if c < 6:
            z.fill = PatternFill("solid", fgColor=GELB)
    ws.cell(row=r, column=5).number_format = EURO
    g = ws.cell(row=r, column=6, value=f'=IF(OR(C{r}="",E{r}=""),"",C{r}*E{r})')
    g.number_format = EURO
    g.fill = PatternFill("solid", fgColor=GRAU)

# Summen (Pflichtangaben 8, 9)
r = letzte + 2
ws.cell(row=r, column=5, value="Nettobetrag").font = Font(name=SCHRIFT, size=11, bold=True)
netto = ws.cell(row=r, column=6, value=f"=SUM(F{erste}:F{letzte})")
netto.number_format = EURO
netto.font = Font(name=SCHRIFT, size=11)
netto.border = RAHMEN
zeile_netto = r

r += 1
ws.cell(
    row=r,
    column=5,
    value='=IF(\'Meine Daten\'!C12="JA","Umsatzsteuer","zzgl. "&\'Meine Daten\'!C13&" % USt.")',
).font = Font(name=SCHRIFT, size=11)
ust = ws.cell(
    row=r,
    column=6,
    value=f'=IF(\'Meine Daten\'!C12="JA",0,F{zeile_netto}*\'Meine Daten\'!C13/100)',
)
ust.number_format = EURO
ust.font = Font(name=SCHRIFT, size=11)
ust.border = RAHMEN
zeile_ust = r

r += 1
ws.cell(row=r, column=5, value="Rechnungsbetrag").font = Font(
    name=SCHRIFT, size=13, bold=True, color="FFFFFF"
)
ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=DUNKEL)
brutto = ws.cell(row=r, column=6, value=f"=F{zeile_netto}+F{zeile_ust}")
brutto.number_format = EURO
brutto.font = Font(name=SCHRIFT, size=13, bold=True, color="FFFFFF")
brutto.fill = PatternFill("solid", fgColor=DUNKEL)
zeile_brutto = r

# Hinweis Steuerbefreiung (Pflichtangabe 9 für Kleinunternehmer)
r += 2
hinweis = ws.cell(
    row=r,
    column=2,
    value='=IF(\'Meine Daten\'!C12="JA",'
    '"Gemäß § 19 UStG wird keine Umsatzsteuer berechnet.","")',
)
hinweis.font = Font(name=SCHRIFT, size=10, italic=True)

# Zahlungsbedingungen und Steuernummer (Pflichtangabe 3)
r += 2
zahlung = ws.cell(
    row=r,
    column=2,
    value='="Zahlbar innerhalb von "&\'Meine Daten\'!C17&" Tagen ohne Abzug auf folgendes Konto:"',
)
zahlung.font = Font(name=SCHRIFT, size=10)

r += 1
ws.cell(
    row=r,
    column=2,
    value='=\'Meine Daten\'!C14&"  |  IBAN "&\'Meine Daten\'!C15&"  |  BIC "&\'Meine Daten\'!C16',
).font = Font(name=SCHRIFT, size=10)

r += 2
ws.cell(
    row=r,
    column=2,
    value='=IF(\'Meine Daten\'!C11<>"","USt-IdNr.: "&\'Meine Daten\'!C11,'
    '"Steuernummer: "&\'Meine Daten\'!C10)',
).font = Font(name=SCHRIFT, size=10)

ws.print_area = f"A1:F{r + 1}"

# ---------------------------------------------------------------------------
# Pflichtangaben — automatische Prüfung
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Pflichtangaben")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 56
ws.column_dimensions["C"].width = 34

titel(ws, "B2", "Prüfung der Pflichtangaben nach § 14 UStG", 16)
ws["B3"] = "Diese Prüfung läuft automatisch. Steht überall „vollständig“, ist die Rechnung formal in Ordnung."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

pruefungen = [
    ("1. Name und Anschrift des Leistenden", "'Meine Daten'!C5", "'Meine Daten'!C7"),
    ("2. Name und Anschrift des Empfängers", "Rechnung!B5", "Rechnung!B7"),
    ("3. Steuernummer oder USt-IdNr.", "'Meine Daten'!C10", "'Meine Daten'!C11"),
    ("4. Ausstellungsdatum", "Rechnung!C13", None),
    ("5. Fortlaufende Rechnungsnummer", "Rechnung!C12", None),
    ("6. Menge und Art der Leistung", f"Rechnung!B{erste}", f"Rechnung!C{erste}"),
    ("7. Zeitpunkt der Leistung", "Rechnung!C14", None),
    ("8. Entgelt (Nettobetrag)", f"Rechnung!F{zeile_netto}", None),
]

r = 5
for beschriftung, feld1, feld2 in pruefungen:
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11)
    b.border = RAHMEN

    if feld2:
        bedingung = f'AND({feld1}<>"",{feld2}<>"")' if "C11" not in feld2 else f'OR({feld1}<>"",{feld2}<>"")'
    else:
        bedingung = f'{feld1}<>""'

    z = ws.cell(row=r, column=3, value=f'=IF({bedingung},"vollständig","FEHLT")')
    z.font = Font(name=SCHRIFT, size=11, bold=True)
    z.border = RAHMEN
    r += 1

# Punkt 9 hängt vom Kleinunternehmerstatus ab
b = ws.cell(row=r, column=2, value="9. Steuersatz und Steuerbetrag / Befreiungshinweis")
b.font = Font(name=SCHRIFT, size=11)
b.border = RAHMEN
z = ws.cell(
    row=r,
    column=3,
    value='=IF(\'Meine Daten\'!C12="JA","vollständig (§ 19 Hinweis gesetzt)",'
    'IF(\'Meine Daten\'!C13>0,"vollständig","FEHLT: Steuersatz eintragen"))',
)
z.font = Font(name=SCHRIFT, size=11, bold=True)
z.border = RAHMEN
letzte_pruefung = r

r += 2
ws.cell(row=r, column=2, value="Gesamtergebnis").font = Font(
    name=SCHRIFT, size=13, bold=True
)
gesamt = ws.cell(
    row=r,
    column=3,
    value=f'=IF(COUNTIF(C5:C{letzte_pruefung},"FEHLT*")>0,'
    f'"NICHT VERSENDEN – es fehlen Angaben","Rechnung ist formal vollständig")',
)
gesamt.font = Font(name=SCHRIFT, size=13, bold=True)
gesamt.fill = PatternFill("solid", fgColor=GELB)
gesamt.border = RAHMEN

r += 3
ws.cell(
    row=r,
    column=2,
    value="Diese Prüfung deckt die Standardfälle ab. Bei Auslandsgeschäften,",
).font = Font(name=SCHRIFT, size=10, color="C00000")
ws.cell(
    row=r + 1,
    column=2,
    value="Reverse Charge oder Differenzbesteuerung gelten zusätzliche Regeln.",
).font = Font(name=SCHRIFT, size=10, color="C00000")

# ---------------------------------------------------------------------------
# Rechnungsbuch
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Rechnungsbuch")
ws.sheet_view.showGridLines = False

titel(ws, "A1", "Rechnungsbuch", 16)
ws["A2"] = "Jede geschriebene Rechnung hier eintragen — so bleiben die Nummern lückenlos."
ws["A2"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

spalten = [
    ("Rechnungsnr.", 16),
    ("Datum", 14),
    ("Kunde", 30),
    ("Betrag (€)", 16),
    ("Bezahlt am", 14),
    ("Status", 22),
    ("Doppelt?", 14),
]
for i, (ueberschrift, breite) in enumerate(spalten, start=1):
    z = ws.cell(row=4, column=i, value=ueberschrift)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.alignment = Alignment(horizontal="center")
    z.border = RAHMEN
    ws.column_dimensions[get_column_letter(i)].width = breite

ws.cell(row=5, column=1, value="2026-001")
ws.cell(row=5, column=2, value="2026-08-05")
ws.cell(row=5, column=3, value="Kunde GmbH")
ws.cell(row=5, column=4, value=1011.50)
for c in range(1, 5):
    ws.cell(row=5, column=c).font = Font(name=SCHRIFT, size=10, italic=True, color="808080")

ende = 4 + BUCH_ZEILEN
for r in range(5, ende + 1):
    for c in range(1, 8):
        z = ws.cell(row=r, column=c)
        z.border = RAHMEN
        z.font = Font(name=SCHRIFT, size=10)
        if c <= 5:
            z.fill = PatternFill("solid", fgColor=GELB)
    ws.cell(row=r, column=2).number_format = DATUM
    ws.cell(row=r, column=4).number_format = EURO
    ws.cell(row=r, column=5).number_format = DATUM

    status = ws.cell(
        row=r,
        column=6,
        value=f'=IF(A{r}="","",IF(E{r}<>"","bezahlt",'
        f'IF(TODAY()-B{r}>\'Meine Daten\'!$C$17,"ÜBERFÄLLIG","offen")))',
    )
    status.fill = PatternFill("solid", fgColor=GRAU)

    doppelt = ws.cell(
        row=r,
        column=7,
        value=f'=IF(A{r}="","",IF(COUNTIF($A$5:$A${ende},A{r})>1,"DOPPELT!",""))',
    )
    doppelt.fill = PatternFill("solid", fgColor=GRAU)

ws.freeze_panes = "A5"

r = ende + 2
ws.cell(row=r, column=3, value="Summe berechnet").font = Font(name=SCHRIFT, size=11, bold=True)
s = ws.cell(row=r, column=4, value=f"=SUM(D5:D{ende})")
s.number_format = EURO
s.font = Font(name=SCHRIFT, size=11, bold=True)
s.fill = PatternFill("solid", fgColor=HELL)
s.border = RAHMEN

ws.cell(row=r + 1, column=3, value="davon noch offen").font = Font(name=SCHRIFT, size=11, bold=True)
o = ws.cell(row=r + 1, column=4, value=f'=SUMIF(E5:E{ende},"",D5:D{ende})')
o.number_format = EURO
o.font = Font(name=SCHRIFT, size=11, bold=True)
o.fill = PatternFill("solid", fgColor=ROT)
o.border = RAHMEN

wb.save("Rechnungsvorlage-Pflichtangaben.xlsx")
print("gespeichert")
