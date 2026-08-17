"""
Erzeugt den Stundensatz- und Preiskalkulator.

Verkaufsprodukt für den Marktplatz. Zielgruppe: Solo-Selbstständige, die nicht
wissen, was sie verlangen müssen, um am Ende des Jahres wirklich von ihrer
Arbeit leben zu können.

Kern des Produkts: Die meisten rechnen "Wunschgehalt geteilt durch 2000 Stunden"
und liegen damit dramatisch zu niedrig. Diese Datei rechnet rückwärts von dem,
was übrig bleiben MUSS — über Steuern, Sozialabgaben, Ausfallzeiten und
unbezahlte Arbeitszeit.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCHRIFT = "Arial"
DUNKEL = "1F3864"
HELL = "D9E2F3"
GELB = "FFF2CC"
GRAU = "F2F2F2"
GRUEN = "E2EFDA"
ROT = "FCE4E4"

EURO = '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'
EURO0 = '#,##0 "€";[Red]-#,##0 "€";"–"'
PROZENT = "0.0%"

rand = Side(style="thin", color="BFBFBF")
RAHMEN = Border(left=rand, right=rand, top=rand, bottom=rand)


def titel(ws, zelle, text, groesse=14, farbe=DUNKEL):
    ws[zelle] = text
    ws[zelle].font = Font(name=SCHRIFT, size=groesse, bold=True, color=farbe)


def block(ws, zeile, text):
    """Abschnittsüberschrift über die Breite."""
    z = ws.cell(row=zeile, column=2, value=text)
    z.font = Font(name=SCHRIFT, size=12, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    for c in (3, 4):
        ws.cell(row=zeile, column=c).fill = PatternFill("solid", fgColor=DUNKEL)


def zeile(ws, r, beschriftung, wert, format_=None, eingabe=False, fett=False,
          hinweis=None, hervorheben=None):
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11, bold=fett)
    b.border = RAHMEN
    b.alignment = Alignment(wrap_text=True, vertical="center")

    w = ws.cell(row=r, column=3, value=wert)
    w.border = RAHMEN
    if format_:
        w.number_format = format_
    if eingabe:
        w.font = Font(name=SCHRIFT, size=11, bold=fett, color="0000FF")
        w.fill = PatternFill("solid", fgColor=GELB)
    else:
        w.font = Font(name=SCHRIFT, size=11, bold=fett)
        if hervorheben:
            w.fill = PatternFill("solid", fgColor=hervorheben)

    if hinweis:
        h = ws.cell(row=r, column=4, value=hinweis)
        h.font = Font(name=SCHRIFT, size=9, color="595959")
        h.alignment = Alignment(wrap_text=True, vertical="center")
    return r + 1


wb = Workbook()

# ---------------------------------------------------------------------------
# Anleitung
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

titel(ws, "B2", "Stundensatz- und Preiskalkulator", 18)

text = [
    ("", None),
    ("Warum die übliche Rechnung falsch ist", "kopf"),
    ("Die meisten rechnen: Wunschgehalt geteilt durch 2.000 Stunden im Jahr.", None),
    ("Das ist der teuerste Rechenfehler in der Selbstständigkeit.", None),
    ("", None),
    ("Warum? Weil du nicht 2.000 Stunden verkaufen kannst. Urlaub, Krankheit,", None),
    ("Feiertage, Angebote schreiben, Buchhaltung, Akquise, Weiterbildung —", None),
    ("all das kostet Zeit, die niemand bezahlt. Realistisch bleiben von 2.000", None),
    ("Stunden etwa 1.000 bis 1.200 fakturierbare übrig.", None),
    ("", None),
    ("Dazu kommen Steuern, Krankenversicherung und Altersvorsorge, die bei", None),
    ("Angestellten der Arbeitgeber mitträgt. Als Selbstständiger zahlst du alles.", None),
    ("", None),
    ("So arbeitest du damit", "kopf"),
    ("1. Blatt „Kalkulation“ öffnen und die gelben Felder ausfüllen.", None),
    ("2. Ganz unten steht dein Mindest-Stundensatz. Das ist die Untergrenze,", None),
    ("   nicht dein Wunschpreis.", None),
    ("3. Blatt „Projektpreis“ nutzen, um einzelne Angebote zu kalkulieren.", None),
    ("4. Blatt „Vergleich“ zeigt, was verschiedene Stundensätze im Jahr bedeuten.", None),
    ("", None),
    ("Was der Kalkulator NICHT kann", "kopf"),
    ("Er sagt dir, was du mindestens verlangen musst — nicht, was der Markt zahlt.", None),
    ("Liegt dein Mindestsatz über dem, was üblich ist, hast du ein Problem, das", None),
    ("sich nicht wegrechnen lässt. Dann musst du entweder effizienter arbeiten,", None),
    ("höherwertige Leistungen anbieten oder die Zielgruppe wechseln.", None),
    ("", None),
    ("Wichtiger Hinweis", "warnung"),
    ("Dieser Kalkulator ist eine Arbeitshilfe, KEINE Steuer- oder Rentenberatung.", None),
    ("Die Steuersätze sind Näherungswerte für eine grobe Orientierung — deine", None),
    ("tatsächliche Steuerlast hängt von Familienstand, weiteren Einkünften und", None),
    ("Freibeträgen ab. Für belastbare Zahlen: Steuerberater fragen.", None),
    ("", None),
    ("Stand: August 2026.", None),
]

r = 3
for inhalt, art in text:
    z = ws.cell(row=r, column=2, value=inhalt)
    if art == "kopf":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color=DUNKEL)
    elif art == "warnung":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color="C00000")
    else:
        z.font = Font(name=SCHRIFT, size=11)
    r += 1

# ---------------------------------------------------------------------------
# Kalkulation
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Kalkulation")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 46
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 52

titel(ws, "B2", "Dein Mindest-Stundensatz", 16)
ws["B3"] = "Gelbe Felder ausfüllen. Alles andere rechnet sich automatisch."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

r = 5
block(ws, r, "1. Was du zum Leben brauchst"); r += 1
r = zeile(ws, r, "Gewünschtes Netto pro Monat", 2500, EURO, eingabe=True,
          hinweis="Was am Monatsende auf deinem Privatkonto ankommen soll.")
zeile_netto_monat = r - 1
r = zeile(ws, r, "Netto pro Jahr", f"=C{zeile_netto_monat}*12", EURO)
zeile_netto_jahr = r - 1

r += 1
block(ws, r, "2. Was zusätzlich abgeht"); r += 1
r = zeile(ws, r, "Einkommensteuer, geschätzt in %", 0.25, PROZENT, eingabe=True,
          hinweis="Grobe Näherung. Bei 30.000–50.000 € Gewinn meist 20–30 %.")
zeile_steuer = r - 1
r = zeile(ws, r, "Krankenversicherung pro Monat", 450, EURO, eingabe=True,
          hinweis="Freiwillig gesetzlich versichert: meist 350–550 €. Privat variiert stark.")
zeile_kv = r - 1
r = zeile(ws, r, "Altersvorsorge pro Monat", 300, EURO, eingabe=True,
          hinweis="Als Selbstständiger zahlt das niemand für dich. Nicht weglassen.")
zeile_av = r - 1

r = zeile(ws, r, "Summe Vorsorge pro Jahr", f"=(C{zeile_kv}+C{zeile_av})*12", EURO)
zeile_vorsorge = r - 1

r = zeile(
    ws, r, "Nötiger Gewinn vor Steuern",
    f"=(C{zeile_netto_jahr}+C{zeile_vorsorge})/(1-C{zeile_steuer})",
    EURO, fett=True, hervorheben=HELL,
    hinweis="Rückwärts gerechnet: Von diesem Betrag bleibt nach Steuern und Vorsorge dein Wunsch-Netto übrig.",
)
zeile_gewinn = r - 1

r += 1
block(ws, r, "3. Deine Betriebskosten"); r += 1
kosten = [
    ("Büro oder Arbeitszimmer pro Monat", 200, "Miete, Strom, Reinigung."),
    ("Telefon, Internet, Mobilfunk", 60, ""),
    ("Software, Werkzeuge, Lizenzen", 80, "Buchhaltung, Fachsoftware, Cloud."),
    ("Versicherungen (betrieblich)", 50, "Betriebshaftpflicht, Rechtsschutz."),
    ("Fahrzeug und Fahrtkosten", 250, "Leasing, Sprit, Wartung, ÖPNV."),
    ("Steuerberater und Buchhaltung", 120, ""),
    ("Weiterbildung und Fachliteratur", 60, ""),
    ("Werbung und Website", 50, ""),
    ("Sonstiges und Puffer", 100, "Was immer dazwischenkommt."),
]
erste_kosten = r
for beschriftung, wert, h in kosten:
    r = zeile(ws, r, beschriftung, wert, EURO, eingabe=True, hinweis=h)
letzte_kosten = r - 1

r = zeile(ws, r, "Betriebskosten pro Jahr",
          f"=SUM(C{erste_kosten}:C{letzte_kosten})*12", EURO, fett=True)
zeile_kosten_jahr = r - 1

r += 1
block(ws, r, "4. Deine verkaufbare Zeit"); r += 1
r = zeile(ws, r, "Arbeitsstunden pro Woche", 40, None, eingabe=True)
zeile_std_woche = r - 1
r = zeile(ws, r, "Arbeitswochen pro Jahr", 52, None, eingabe=True)
zeile_wochen = r - 1
r = zeile(ws, r, "Urlaubswochen", 6, None, eingabe=True,
          hinweis="Nimm dir echten Urlaub. Wer 0 einträgt, kalkuliert sich in den Burnout.")
zeile_urlaub = r - 1
r = zeile(ws, r, "Krankheitswochen", 2, None, eingabe=True,
          hinweis="Auch Selbstständige werden krank. Realistisch: 2 Wochen.")
zeile_krank = r - 1
r = zeile(ws, r, "Feiertage in Wochen", 2, None, eingabe=True)
zeile_feiertage = r - 1

r = zeile(ws, r, "Arbeitswochen effektiv",
          f"=C{zeile_wochen}-C{zeile_urlaub}-C{zeile_krank}-C{zeile_feiertage}")
zeile_wochen_netto = r - 1

r = zeile(ws, r, "Bruttoarbeitsstunden pro Jahr",
          f"=C{zeile_wochen_netto}*C{zeile_std_woche}")
zeile_std_brutto = r - 1

r = zeile(ws, r, "Anteil unbezahlte Arbeit in %", 0.35, PROZENT, eingabe=True,
          hinweis="Angebote, Buchhaltung, Akquise, Weiterbildung, E-Mails. Realistisch 30–45 %.")
zeile_unbezahlt = r - 1

r = zeile(ws, r, "Verkaufbare Stunden pro Jahr",
          f"=C{zeile_std_brutto}*(1-C{zeile_unbezahlt})",
          "#,##0", fett=True, hervorheben=HELL,
          hinweis="Nur diese Stunden kannst du in Rechnung stellen.")
zeile_std_verkaufbar = r - 1

r += 1
block(ws, r, "5. Ergebnis"); r += 1
r = zeile(ws, r, "Nötiger Jahresumsatz",
          f"=C{zeile_gewinn}+C{zeile_kosten_jahr}", EURO, fett=True)
zeile_umsatz = r - 1

r += 1
b = ws.cell(row=r, column=2, value="MINDEST-STUNDENSATZ (netto)")
b.font = Font(name=SCHRIFT, size=14, bold=True, color="FFFFFF")
b.fill = PatternFill("solid", fgColor=DUNKEL)
b.border = RAHMEN
s = ws.cell(row=r, column=3,
            value=f"=IF(C{zeile_std_verkaufbar}<=0,\"\",C{zeile_umsatz}/C{zeile_std_verkaufbar})")
s.font = Font(name=SCHRIFT, size=14, bold=True, color="FFFFFF")
s.fill = PatternFill("solid", fgColor=DUNKEL)
s.number_format = EURO
s.border = RAHMEN
h = ws.cell(row=r, column=4,
            value="Das ist die Untergrenze, nicht dein Wunschpreis. Darunter arbeitest du mit Verlust.")
h.font = Font(name=SCHRIFT, size=9, bold=True, color="C00000")
h.alignment = Alignment(wrap_text=True, vertical="center")
zeile_satz = r

r += 2
r = zeile(ws, r, "Tagessatz (8 Stunden)", f"=C{zeile_satz}*8", EURO, fett=True)
r = zeile(ws, r, "Empfohlener Angebotssatz (+20 % Puffer)",
          f"=C{zeile_satz}*1.2", EURO, fett=True, hervorheben=GRUEN,
          hinweis="Der Mindestsatz lässt keinen Raum für Rabatte, Ausfälle oder Investitionen.")

r += 1
ws.cell(row=r, column=2,
        value="Näherungsrechnung, keine Steuerberatung. Steuersatz ist geschätzt.").font = Font(
    name=SCHRIFT, size=9, italic=True, color="C00000")

# ---------------------------------------------------------------------------
# Projektpreis
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Projektpreis")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 18

titel(ws, "B2", "Einzelnes Projekt kalkulieren", 16)
ws["B3"] = "Für Festpreis-Angebote. Der Stundensatz kommt automatisch aus dem Blatt „Kalkulation“."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

for i, (ueberschrift, breite) in enumerate(
    zip(["Arbeitspaket", "Stunden", "Stundensatz", "Summe"], [40, 16, 16, 18]), start=2
):
    z = ws.cell(row=5, column=i, value=ueberschrift)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.alignment = Alignment(horizontal="center")
    z.border = RAHMEN

ws.cell(row=6, column=2, value="Beispiel: Konzept und Abstimmung")
ws.cell(row=6, column=3, value=8)
for c in (2, 3):
    ws.cell(row=6, column=c).font = Font(name=SCHRIFT, size=10, italic=True, color="808080")

for r in range(6, 26):
    for c in range(2, 6):
        z = ws.cell(row=r, column=c)
        z.border = RAHMEN
        z.font = Font(name=SCHRIFT, size=10)
        if c <= 3:
            z.fill = PatternFill("solid", fgColor=GELB)
    satz = ws.cell(row=r, column=4, value=f"=Kalkulation!$C${zeile_satz}")
    satz.number_format = EURO
    satz.fill = PatternFill("solid", fgColor=GRAU)
    summe = ws.cell(row=r, column=5, value=f'=IF(C{r}="","",C{r}*D{r})')
    summe.number_format = EURO
    summe.fill = PatternFill("solid", fgColor=GRAU)

r = 27
ws.cell(row=r, column=4, value="Summe Arbeitszeit").font = Font(name=SCHRIFT, size=11, bold=True)
s = ws.cell(row=r, column=5, value="=SUM(E6:E25)")
s.number_format = EURO
s.font = Font(name=SCHRIFT, size=11, bold=True)
s.border = RAHMEN
zeile_arbeitszeit = r

r += 1
ws.cell(row=r, column=4, value="Materialkosten").font = Font(name=SCHRIFT, size=11)
m = ws.cell(row=r, column=5, value=0)
m.number_format = EURO
m.font = Font(name=SCHRIFT, size=11, color="0000FF")
m.fill = PatternFill("solid", fgColor=GELB)
m.border = RAHMEN
zeile_material = r

r += 1
ws.cell(row=r, column=4, value="Risikoaufschlag in %").font = Font(name=SCHRIFT, size=11)
risiko = ws.cell(row=r, column=5, value=0.15)
risiko.number_format = PROZENT
risiko.font = Font(name=SCHRIFT, size=11, color="0000FF")
risiko.fill = PatternFill("solid", fgColor=GELB)
risiko.border = RAHMEN
zeile_risiko = r

r += 1
ws.cell(row=r, column=4, value="ANGEBOTSPREIS (netto)").font = Font(
    name=SCHRIFT, size=13, bold=True, color="FFFFFF")
ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=DUNKEL)
p = ws.cell(row=r, column=5,
            value=f"=(E{zeile_arbeitszeit}+E{zeile_material})*(1+E{zeile_risiko})")
p.number_format = EURO
p.font = Font(name=SCHRIFT, size=13, bold=True, color="FFFFFF")
p.fill = PatternFill("solid", fgColor=DUNKEL)
p.border = RAHMEN

r += 2
ws.cell(row=r, column=2,
        value="Der Risikoaufschlag deckt ab, dass Projekte fast immer länger dauern als geschätzt.").font = Font(
    name=SCHRIFT, size=9, italic=True, color="595959")

# ---------------------------------------------------------------------------
# Vergleich
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Vergleich")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for spalte, breite in zip("BCDEF", [22, 24, 24, 24, 24]):
    ws.column_dimensions[spalte].width = breite

titel(ws, "B2", "Was verschiedene Stundensätze bedeuten", 16)
ws["B3"] = "Basis: deine verkaufbaren Stunden und Kosten aus dem Blatt „Kalkulation“."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

saetze = [40, 60, 80, 100, 120]
z = ws.cell(row=5, column=2, value="Stundensatz")
z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
z.fill = PatternFill("solid", fgColor=DUNKEL)
z.border = RAHMEN

for i, satz in enumerate(saetze, start=3):
    z = ws.cell(row=5, column=i, value=satz)
    z.font = Font(name=SCHRIFT, size=11, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.number_format = EURO0
    z.alignment = Alignment(horizontal="center")
    z.border = RAHMEN

zeilen_vergleich = [
    ("Jahresumsatz", f"=C$5*Kalkulation!$C${zeile_std_verkaufbar}", EURO0),
    ("minus Betriebskosten", f"=-Kalkulation!$C${zeile_kosten_jahr}", EURO0),
    ("Gewinn vor Steuern", None, EURO0),
    ("minus Steuern (geschätzt)", None, EURO0),
    ("minus Vorsorge", f"=-Kalkulation!$C${zeile_vorsorge}", EURO0),
    ("Netto pro Jahr", None, EURO0),
    ("Netto pro Monat", None, EURO),
]

r = 6
for i, (beschriftung, formel, fmt) in enumerate(zeilen_vergleich):
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=10, bold=(beschriftung.startswith("Netto")))
    b.border = RAHMEN

    for c in range(3, 8):
        sp = get_column_letter(c)
        if beschriftung == "Jahresumsatz":
            f = f"={sp}$5*Kalkulation!$C${zeile_std_verkaufbar}"
        elif beschriftung == "minus Betriebskosten":
            f = f"=-Kalkulation!$C${zeile_kosten_jahr}"
        elif beschriftung == "Gewinn vor Steuern":
            f = f"={sp}6+{sp}7"
        elif beschriftung == "minus Steuern (geschätzt)":
            f = f"=-MAX(0,{sp}8)*Kalkulation!$C${zeile_steuer}"
        elif beschriftung == "minus Vorsorge":
            f = f"=-Kalkulation!$C${zeile_vorsorge}"
        elif beschriftung == "Netto pro Jahr":
            f = f"={sp}8+{sp}9+{sp}10"
        else:
            f = f"={sp}11/12"

        z = ws.cell(row=r, column=c, value=f)
        z.number_format = fmt
        z.border = RAHMEN
        z.font = Font(name=SCHRIFT, size=10, bold=beschriftung.startswith("Netto"))
        if beschriftung.startswith("Netto"):
            z.fill = PatternFill("solid", fgColor=GRUEN)
    r += 1

r += 1
ws.cell(row=r, column=2,
        value="Rote Werte bedeuten: Bei diesem Stundensatz machst du Verlust.").font = Font(
    name=SCHRIFT, size=10, color="C00000")
ws.cell(row=r + 1, column=2,
        value="Näherungsrechnung mit geschätztem Steuersatz — keine Steuerberatung.").font = Font(
    name=SCHRIFT, size=9, italic=True, color="595959")

wb.save("Stundensatz-Kalkulator.xlsx")
print(f"gespeichert (Stundensatz in Kalkulation!C{zeile_satz})")
