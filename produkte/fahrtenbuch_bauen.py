"""
Erzeugt das finanzamtskonforme Fahrtenbuch.

Verkaufsprodukt. Zielgruppe: Selbstständige mit Firmenwagen oder privatem Auto
im Betriebsvermögen, die die 1-%-Regelung vermeiden wollen.

Kern: Ein Fahrtenbuch wird vom Finanzamt nur anerkannt, wenn es zeitnah,
lückenlos und in geschlossener Form geführt wird. Diese Vorlage erzwingt die
Pflichtangaben und rechnet den Privatanteil automatisch aus.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCHRIFT = "Arial"
DUNKEL = "1F3864"
HELL = "D9E2F3"
GELB = "FFF2CC"
GRAU = "F2F2F2"
GRUEN = "E2EFDA"
ROT = "FCE4E4"

EURO = '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'
KM = '#,##0 "km"'
DATUM = "TT.MM.JJJJ"
PROZENT = "0.0%"

rand = Side(style="thin", color="BFBFBF")
RAHMEN = Border(left=rand, right=rand, top=rand, bottom=rand)

ZEILEN = 400
ARTEN = ["betrieblich", "Fahrt Wohnung–Betrieb", "privat"]


def titel(ws, zelle, text, groesse=14, farbe=DUNKEL):
    ws[zelle] = text
    ws[zelle].font = Font(name=SCHRIFT, size=groesse, bold=True, color=farbe)


wb = Workbook()

# --------------------------------------------------------------------------
# Anleitung
# --------------------------------------------------------------------------
ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

titel(ws, "B2", "Fahrtenbuch für Selbstständige", 18)

text = [
    ("", None),
    ("Warum ein Fahrtenbuch?", "kopf"),
    ("Nutzt du ein Fahrzeug aus dem Betriebsvermögen auch privat, musst du den", None),
    ("privaten Anteil versteuern. Dafür gibt es zwei Wege:", None),
    ("", None),
    ("   1-%-Regelung: Pauschal 1 % des Bruttolistenpreises pro Monat.", None),
    ("   Einfach, aber oft teuer — besonders bei teuren Fahrzeugen und", None),
    ("   wenig Privatnutzung.", None),
    ("", None),
    ("   Fahrtenbuch: Du weist den tatsächlichen Privatanteil nach und", None),
    ("   versteuerst nur diesen. Mehr Aufwand, aber meist deutlich günstiger.", None),
    ("", None),
    ("Diese Vorlage rechnet dir beides aus und zeigt, welcher Weg für dich", None),
    ("günstiger ist.", None),
    ("", None),
    ("Die vier Anforderungen des Finanzamts", "kopf"),
    ("Ein Fahrtenbuch wird nur anerkannt, wenn es alle vier erfüllt:", None),
    ("", None),
    ("   1. ZEITNAH geführt — also direkt nach der Fahrt, nicht am Jahresende.", None),
    ("   2. LÜCKENLOS — jede Fahrt, keine Lücken im Kilometerstand.", None),
    ("   3. GESCHLOSSEN — nachträglich nicht änderbar.", None),
    ("   4. VOLLSTÄNDIG — mit allen Pflichtangaben je Fahrt.", None),
    ("", None),
    ("Pflichtangaben je betrieblicher Fahrt", "kopf"),
    ("   • Datum", None),
    ("   • Kilometerstand bei Beginn und Ende", None),
    ("   • Reiseziel (Ort und Straße)", None),
    ("   • Reisezweck", None),
    ("   • Aufgesuchter Geschäftspartner oder Kunde", None),
    ("", None),
    ("Bei Privatfahrten genügt die Angabe der gefahrenen Kilometer.", None),
    ("Bei Fahrten Wohnung–Betrieb genügt ein entsprechender Vermerk.", None),
    ("", None),
    ("Der wichtigste Punkt", "warnung"),
    ("Eine Excel-Datei gilt für sich genommen NICHT als geschlossene Form —", None),
    ("sie lässt sich jederzeit ändern. Damit das Fahrtenbuch anerkannt wird,", None),
    ("musst du es regelmäßig (mindestens monatlich) ausdrucken oder als PDF", None),
    ("sichern und diese Fassung unverändert aufbewahren.", None),
    ("", None),
    ("Nutze diese Datei als Erfassungshilfe — der Nachweis ist der monatliche", None),
    ("Ausdruck. Wer das nicht macht, riskiert, dass das Finanzamt das gesamte", None),
    ("Fahrtenbuch verwirft und die 1-%-Regelung ansetzt. Rückwirkend.", None),
    ("", None),
    ("Hinweis", "kopf"),
    ("Diese Vorlage ist eine Arbeitshilfe, keine Steuerberatung. Die Anerkennung", None),
    ("entscheidet im Einzelfall das Finanzamt. Stand: August 2026.", None),
]

r = 3
for inhalt, art in text:
    z = ws.cell(row=r, column=2, value=inhalt)
    if art == "kopf":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color=DUNKEL)
    elif art == "warnung":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color="C00000")
    else:
        z.font = Font(name=SCHRIFT, size=11)
    r += 1

# --------------------------------------------------------------------------
# Fahrzeug
# --------------------------------------------------------------------------
ws = wb.create_sheet("Fahrzeug")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 46

titel(ws, "B2", "Fahrzeugdaten", 16)
ws["B3"] = "Einmalig ausfüllen."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

felder = [
    ("Fahrzeug (Marke, Modell)", "VW Caddy", ""),
    ("Amtliches Kennzeichen", "KN-XY 123", ""),
    ("Bruttolistenpreis bei Erstzulassung", 32000, "Neupreis laut Liste, auch bei Gebrauchtkauf. Basis für die 1-%-Regelung."),
    ("Kilometerstand am 1.1.", 15000, ""),
    ("Wirtschaftsjahr", 2026, ""),
    ("Entfernung Wohnung–Betrieb (km)", 12, "Einfache Strecke. Für die Pendlerpauschale."),
    ("Gesamtkosten Fahrzeug pro Jahr", 6800, "Abschreibung, Sprit, Versicherung, Wartung, Steuer."),
    ("Persönlicher Steuersatz", 0.30, "Grobe Näherung für den Vergleich."),
]

r = 5
for beschriftung, wert, h in felder:
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11, bold=True)
    b.border = RAHMEN
    w = ws.cell(row=r, column=3, value=wert)
    w.font = Font(name=SCHRIFT, size=11, color="0000FF")
    w.fill = PatternFill("solid", fgColor=GELB)
    w.border = RAHMEN
    if "preis" in beschriftung.lower() or "kosten" in beschriftung.lower():
        w.number_format = EURO
    if "Steuersatz" in beschriftung:
        w.number_format = PROZENT
    if h:
        hz = ws.cell(row=r, column=4, value=h)
        hz.font = Font(name=SCHRIFT, size=9, color="595959")
        hz.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

ZEILE_LISTENPREIS = 7
ZEILE_KM_START = 8
ZEILE_ENTFERNUNG = 10
ZEILE_KOSTEN = 11
ZEILE_STEUERSATZ = 12

# --------------------------------------------------------------------------
# Fahrtenbuch
# --------------------------------------------------------------------------
ws = wb.create_sheet("Fahrtenbuch")
ws.sheet_view.showGridLines = False

titel(ws, "A1", "Fahrtenbuch", 16)
ws["A2"] = "Jede Fahrt sofort eintragen. Am Monatsende ausdrucken und aufbewahren."
ws["A2"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

spalten = [
    ("Datum", 13),
    ("km Beginn", 12),
    ("km Ende", 12),
    ("gefahren", 12),
    ("Art der Fahrt", 22),
    ("Reiseziel (Ort, Straße)", 32),
    ("Reisezweck", 30),
    ("Geschäftspartner / Kunde", 28),
    ("Prüfung", 26),
]
for i, (ueberschrift, breite) in enumerate(spalten, start=1):
    z = ws.cell(row=4, column=i, value=ueberschrift)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.alignment = Alignment(horizontal="center", wrap_text=True)
    z.border = RAHMEN
    ws.column_dimensions[get_column_letter(i)].width = breite
ws.row_dimensions[4].height = 28

# Beispielzeile
beispiel = ["2026-01-08", 15000, 15042, None, "betrieblich",
            "Singen, Industriestr. 4", "Standortbesichtigung", "Muster GmbH"]
for i, w in enumerate(beispiel, start=1):
    if w is None:
        continue
    z = ws.cell(row=5, column=i, value=w)
    z.font = Font(name=SCHRIFT, size=9, italic=True, color="808080")

erste, letzte = 5, 4 + ZEILEN
for r in range(erste, letzte + 1):
    for c in range(1, 10):
        z = ws.cell(row=r, column=c)
        z.border = RAHMEN
        z.font = Font(name=SCHRIFT, size=9)
        if c in (1, 2, 3, 5, 6, 7, 8):
            z.fill = PatternFill("solid", fgColor=GELB)
    ws.cell(row=r, column=1).number_format = DATUM

    gefahren = ws.cell(row=r, column=4,
                       value=f'=IF(OR(B{r}="",C{r}=""),"",C{r}-B{r})')
    gefahren.number_format = KM
    gefahren.fill = PatternFill("solid", fgColor=GRAU)

    # Prüfspalte: erkennt die drei häufigsten Fehler
    pruefung = ws.cell(
        row=r, column=9,
        value=(
            f'=IF(A{r}="","",'
            f'IF(C{r}<B{r},"km Ende < km Beginn",'
            f'IF(AND(R OW()>{erste},B{r}<>"",INDEX($C${erste}:$C${letzte},ROW()-{erste})<>"",'
            f'B{r}<>INDEX($C${erste}:$C${letzte},ROW()-{erste})),"LÜCKE im km-Stand",'
            f'IF(AND(E{r}="betrieblich",OR(F{r}="",G{r}="",H{r}="")),"Pflichtangabe fehlt",'
            f'"ok"))))'
        ).replace("R OW", "ROW"),
    )
    pruefung.fill = PatternFill("solid", fgColor=GRAU)
    pruefung.font = Font(name=SCHRIFT, size=9, bold=True)

pruef_art = DataValidation(type="list", formula1='"' + ",".join(ARTEN) + '"',
                           allow_blank=True, showDropDown=False)
pruef_art.error = "Bitte betrieblich, Fahrt Wohnung–Betrieb oder privat wählen."
pruef_art.errorTitle = "Unbekannte Fahrtart"
ws.add_data_validation(pruef_art)
pruef_art.add(f"E{erste}:E{letzte}")

ws.freeze_panes = "A5"

# --------------------------------------------------------------------------
# Auswertung
# --------------------------------------------------------------------------
ws = wb.create_sheet("Auswertung")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 46
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 48

titel(ws, "B2", "Auswertung", 16)

def zeile(r, beschriftung, wert, fmt=None, fett=False, farbe=None, hinweis=None):
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11, bold=fett)
    b.border = RAHMEN
    w = ws.cell(row=r, column=3, value=wert)
    w.font = Font(name=SCHRIFT, size=11, bold=fett)
    w.border = RAHMEN
    if fmt:
        w.number_format = fmt
    if farbe:
        w.fill = PatternFill("solid", fgColor=farbe)
    if hinweis:
        h = ws.cell(row=r, column=4, value=hinweis)
        h.font = Font(name=SCHRIFT, size=9, color="595959")
        h.alignment = Alignment(wrap_text=True, vertical="center")
    return r + 1

r = 5
z = ws.cell(row=r, column=2, value="Gefahrene Kilometer")
z.font = Font(name=SCHRIFT, size=12, bold=True, color="FFFFFF")
z.fill = PatternFill("solid", fgColor=DUNKEL)
for c in (3, 4):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=DUNKEL)
r += 1

r = zeile(r, "betrieblich",
          f'=SUMIF(Fahrtenbuch!$E${erste}:$E${letzte},"betrieblich",Fahrtenbuch!$D${erste}:$D${letzte})', KM)
Z_BETRIEB = r - 1
r = zeile(r, "Fahrten Wohnung–Betrieb",
          f'=SUMIF(Fahrtenbuch!$E${erste}:$E${letzte},"Fahrt Wohnung–Betrieb",Fahrtenbuch!$D${erste}:$D${letzte})', KM)
Z_PENDEL = r - 1
r = zeile(r, "privat",
          f'=SUMIF(Fahrtenbuch!$E${erste}:$E${letzte},"privat",Fahrtenbuch!$D${erste}:$D${letzte})', KM)
Z_PRIVAT = r - 1
r = zeile(r, "Gesamt", f"=C{Z_BETRIEB}+C{Z_PENDEL}+C{Z_PRIVAT}", KM, fett=True, farbe=HELL)
Z_GESAMT = r - 1

r = zeile(r, "Privatanteil in Prozent",
          f'=IF(C{Z_GESAMT}=0,0,C{Z_PRIVAT}/C{Z_GESAMT})', PROZENT, fett=True,
          hinweis="Nur dieser Anteil der Fahrzeugkosten ist privat und muss versteuert werden.")
Z_ANTEIL = r - 1

r += 1
z = ws.cell(row=r, column=2, value="Vergleich: Fahrtenbuch oder 1-%-Regelung")
z.font = Font(name=SCHRIFT, size=12, bold=True, color="FFFFFF")
z.fill = PatternFill("solid", fgColor=DUNKEL)
for c in (3, 4):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=DUNKEL)
r += 1

r = zeile(r, "Geldwerter Vorteil laut Fahrtenbuch",
          f"=Fahrzeug!C{ZEILE_KOSTEN}*C{Z_ANTEIL}", EURO,
          hinweis="Tatsächliche Kosten mal Privatanteil.")
Z_FB = r - 1

r = zeile(r, "Geldwerter Vorteil laut 1-%-Regelung",
          f"=Fahrzeug!C{ZEILE_LISTENPREIS}*0.01*12", EURO,
          hinweis="1 % des Bruttolistenpreises pro Monat.")
Z_EINPROZENT = r - 1

r = zeile(r, "Unterschied",
          f"=C{Z_EINPROZENT}-C{Z_FB}", EURO, fett=True,
          hinweis="Positiv heißt: Das Fahrtenbuch spart dir diesen Betrag an zu versteuerndem Vorteil.")
Z_DIFF = r - 1

r = zeile(r, "Steuerersparnis geschätzt",
          f"=MAX(0,C{Z_DIFF})*Fahrzeug!C{ZEILE_STEUERSATZ}", EURO, fett=True, farbe=GRUEN,
          hinweis="Grobe Näherung mit deinem Steuersatz.")

r += 1
b = ws.cell(row=r, column=2, value="EMPFEHLUNG")
b.font = Font(name=SCHRIFT, size=13, bold=True, color="FFFFFF")
b.fill = PatternFill("solid", fgColor=DUNKEL)
b.border = RAHMEN
e = ws.cell(
    row=r, column=3,
    value=f'=IF(C{Z_GESAMT}=0,"Noch keine Fahrten erfasst",'
          f'IF(C{Z_DIFF}>0,"Fahrtenbuch lohnt sich","1-%-Regelung ist günstiger"))',
)
e.font = Font(name=SCHRIFT, size=12, bold=True)
e.fill = PatternFill("solid", fgColor=GELB)
e.border = RAHMEN

r += 2
z = ws.cell(row=r, column=2, value="Kontrolle der Vollständigkeit")
z.font = Font(name=SCHRIFT, size=12, bold=True, color="FFFFFF")
z.fill = PatternFill("solid", fgColor=DUNKEL)
for c in (3, 4):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=DUNKEL)
r += 1

r = zeile(r, "Fahrten erfasst",
          f'=COUNTIF(Fahrtenbuch!$A${erste}:$A${letzte},"<>")', "#,##0")
r = zeile(r, "Beanstandete Zeilen",
          f'=COUNTIFS(Fahrtenbuch!$I${erste}:$I${letzte},"<>ok",'
          f'Fahrtenbuch!$I${erste}:$I${letzte},"<>")', "#,##0", fett=True, farbe=ROT,
          hinweis="Muss 0 sein. Die Spalte „Prüfung“ im Fahrtenbuch zeigt, was fehlt.")

r += 1
ws.cell(row=r, column=2,
        value="Denk daran: monatlich ausdrucken oder als PDF sichern.").font = Font(
    name=SCHRIFT, size=10, bold=True, color="C00000")
ws.cell(row=r + 1, column=2,
        value="Eine Excel-Datei allein erfüllt die geschlossene Form nicht.").font = Font(
    name=SCHRIFT, size=10, color="C00000")
ws.cell(row=r + 3, column=2,
        value="Arbeitshilfe, keine Steuerberatung. Stand: August 2026.").font = Font(
    name=SCHRIFT, size=9, italic=True, color="595959")

wb.save("Fahrtenbuch-finanzamtskonform.xlsx")
print("gespeichert")
