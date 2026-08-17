"""
Erzeugt den Rentabilitätsrechner für Automatenaufsteller.

Verkaufsprodukt für den Marktplatz. Zielgruppe: Leute, die einen Snack- oder
Getränkeautomaten aufstellen wollen und wissen müssen, ob sich ein konkreter
Standort trägt — bevor sie den Vertrag unterschreiben.

Kern des Produkts: Die meisten rechnen "Automat kostet 3.000 €, ich verdiene
50 Cent pro Verkauf, also nach 6.000 Verkäufen bin ich raus". Dabei fehlen
Standortmiete, Strom, die eigenen Befüllfahrten, Schwund und Reparaturen. Diese
Datei rechnet vollständig — und sagt am Ende eine Zahl: die Verkäufe pro Tag,
ab denen der Standort Geld verdient statt kostet.

Aufbau:
  Anleitung           — was wo eingetragen wird
  Standort-Rechner    — ein Standort, vollständige Rechnung, Break-even
  Standortvergleich   — bis zu 8 Standorte nebeneinander
  Produktkalkulation  — Einkaufspreis, Verkaufspreis, Marge je Artikel
  Nachschlagen        — Erfahrungswerte und Fußangeln

Zu den Formeln: bewusst nur Funktionen aus der Excel-97-Ära (SUM, IF, IFERROR,
SUMIFS, INDEX, MATCH, ROUND). Alles Neuere läuft in LibreOffice, Numbers oder
Google Tabellen Gefahr, als #NAME? anzukommen — und der Käufer sieht dann eine
kaputte Datei.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCHRIFT = "Arial"
DUNKEL = "1F3864"
HELL = "D9E2F3"
GELB = "FFF2CC"
GRAU = "F2F2F2"
GRUEN = "E2EFDA"
ROT = "FCE4E4"

EURO = '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'
EURO0 = '#,##0 "€";[Red]-#,##0 "€";"–"'
PROZENT = "0.0%"
ZAHL1 = '#,##0.0;[Red]-#,##0.0;"–"'
ZAHL0 = '#,##0;[Red]-#,##0;"–"'

rand = Side(style="thin", color="BFBFBF")
RAHMEN = Border(left=rand, right=rand, top=rand, bottom=rand)


# ---------------------------------------------------------------------------
# Bausteine
# ---------------------------------------------------------------------------

def titel(ws, zelle, text, groesse=14, farbe=DUNKEL):
    ws[zelle] = text
    ws[zelle].font = Font(name=SCHRIFT, size=groesse, bold=True, color=farbe)


def block(ws, zeile, text, bis_spalte=4):
    z = ws.cell(row=zeile, column=2, value=text)
    z.font = Font(name=SCHRIFT, size=12, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    for c in range(3, bis_spalte + 1):
        ws.cell(row=zeile, column=c).fill = PatternFill("solid", fgColor=DUNKEL)


def zeile(ws, r, beschriftung, wert, format_=None, eingabe=False, fett=False,
          hinweis=None, hervorheben=None):
    """Eine Rechenzeile: Beschriftung | Wert | Erläuterung."""
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11, bold=fett)
    b.border = RAHMEN
    b.alignment = Alignment(wrap_text=True, vertical="center")

    w = ws.cell(row=r, column=3, value=wert)
    w.border = RAHMEN
    if format_:
        w.number_format = format_
    if eingabe:
        # Blau auf Gelb: die durchgehende Kennzeichnung für "hier tippst du".
        w.font = Font(name=SCHRIFT, size=11, bold=fett, color="0000FF")
        w.fill = PatternFill("solid", fgColor=GELB)
    else:
        w.font = Font(name=SCHRIFT, size=11, bold=fett)
        if hervorheben:
            w.fill = PatternFill("solid", fgColor=hervorheben)

    if hinweis:
        h = ws.cell(row=r, column=4, value=hinweis)
        h.font = Font(name=SCHRIFT, size=9, color="595959")
        h.alignment = Alignment(wrap_text=True, vertical="center")
    return r + 1


def text(ws, r, inhalt, groesse=10, fett=False, farbe="000000", spalte=2):
    z = ws.cell(row=r, column=spalte, value=inhalt)
    z.font = Font(name=SCHRIFT, size=groesse, bold=fett, color=farbe)
    z.alignment = Alignment(wrap_text=True, vertical="top")
    return r + 1


def breiten(ws, paare):
    for spalte, breite in paare.items():
        ws.column_dimensions[spalte].width = breite


wb = Workbook()


# ===========================================================================
# 1 — Anleitung
# ===========================================================================

ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
breiten(ws, {"A": 2, "B": 96})

r = 2
titel(ws, "B2", "Rentabilitätsrechner für Automatenaufsteller", 18)
r = 4
r = text(ws, r, "Diese Datei beantwortet eine einzige Frage: Trägt sich dieser Standort?", 12, True)
r += 1
r = text(ws, r,
         "Sie rechnet nicht schön, sondern vollständig. Enthalten sind auch die Posten, die in "
         "den meisten Überschlagsrechnungen fehlen — die eigenen Befüllfahrten, der Strom, der "
         "Schwund und die Rücklage für Reparaturen. Am Ende steht eine einzige Zahl: wie viele "
         "Verkäufe pro Tag der Automat schaffen muss, damit er Geld verdient statt kostet.")
r += 1

r = text(ws, r, "SO GEHST DU VOR", 11, True, DUNKEL)
r += 1
for schritt in [
    "1.  Blatt „Produktkalkulation“ — trag ein, was du einkaufst und was du verlangen willst. "
    "Das ergibt deinen Rohertrag je Verkauf.",
    "2.  Blatt „Standort-Rechner“ — trag die Angaben zum konkreten Standort ein. Alle gelben "
    "Felder ausfüllen, die blauen Zahlen sind deine Eingaben.",
    "3.  Lies unten das Ergebnis: Gewinn pro Monat, Break-even in Verkäufen pro Tag und die "
    "Amortisationsdauer des Automaten.",
    "4.  Blatt „Standortvergleich“ — trag mehrere Standorte nebeneinander ein und entscheide, "
    "welchen du zuerst angehst.",
]:
    r = text(ws, r, schritt)
r += 1

r = text(ws, r, "DIE FARBEN", 11, True, DUNKEL)
r += 1
r = text(ws, r, "Gelb hinterlegt mit blauer Schrift  =  hier trägst du ein.")
r = text(ws, r, "Weiß mit schwarzer Schrift  =  rechnet die Datei aus. Nicht überschreiben, "
                "sonst ist die Formel weg.")
r = text(ws, r, "Grün oder rot hinterlegt  =  Ergebnis, auf das es ankommt.")
r += 1

r = text(ws, r, "DREI FEHLER, DIE RICHTIG GELD KOSTEN", 11, True, DUNKEL)
r += 1
r = text(ws, r,
         "•  Die eigene Arbeitszeit nicht mitrechnen. Wer für 40 € Rohertrag im Monat zweimal "
         "30 km fährt und je eine Stunde befüllt, arbeitet unter Mindestlohn. Diese Datei "
         "rechnet deine Zeit mit einem Stundensatz, den du selbst festlegst.")
r = text(ws, r,
         "•  Umsatzbeteiligung als kleine Nebensache behandeln. 15 % vom Bruttoumsatz können "
         "mehr sein als deine gesamte Marge, wenn du günstig verkaufst. Die Datei zeigt beide "
         "Modelle — feste Miete und Beteiligung — nebeneinander.")
r = text(ws, r,
         "•  Keine Rücklage für Reparaturen bilden. Ein Kühlaggregat oder ein Münzprüfer kostet "
         "im Ernstfall mehrere hundert Euro. Wer damit nicht rechnet, hält den Standort für "
         "rentabel, bis der erste Schaden kommt.")
r += 1

r = text(ws, r, "ZUR UMSATZSTEUER", 11, True, DUNKEL)
r += 1
r = text(ws, r,
         "Bei Automatenverkäufen gilt nicht durchgängig ein Steuersatz. Für Speisen und für "
         "Getränke gelten unterschiedliche Sätze, und die Abgrenzung hat sich in den letzten "
         "Jahren mehrfach geändert. Deshalb trägst du den Steuersatz in der Produktkalkulation "
         "je Artikel selbst ein — voreingestellt sind 7 % für Lebensmittel und 19 % für "
         "Getränke.")
r = text(ws, r,
         "Kläre die Einordnung deiner Artikel mit deinem Steuerberater. Als Kleinunternehmer "
         "nach § 19 UStG weist du gar keine Umsatzsteuer aus — dann setzt du in der "
         "Produktkalkulation überall 0 % ein.", farbe="C00000")
r += 2

r = text(ws, r,
         "Arbeitshilfe, keine Steuer- oder Rechtsberatung. Zahlen sind Erfahrungswerte und "
         "müssen für den eigenen Fall geprüft werden. Stand: August 2026.", 9, farbe="808080")


# ===========================================================================
# 2 — Produktkalkulation
#
# Muss vor dem Standort-Rechner kommen: der Rohertrag je Verkauf ist dessen
# wichtigste Eingangsgröße.
# ===========================================================================

pk = wb.create_sheet("Produktkalkulation")
pk.sheet_view.showGridLines = False
breiten(pk, {"A": 2, "B": 30, "C": 13, "D": 11, "E": 13, "F": 12, "G": 13,
             "H": 13, "I": 12, "J": 34})

titel(pk, "B2", "Produktkalkulation", 16)
pk["B3"] = ("Was kaufst du ein, was verlangst du, was bleibt hängen? Der Durchschnitt unten "
            "geht in den Standort-Rechner ein.")
pk["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

KOPF = ["Artikel", "Einkauf netto", "USt-Satz", "Verkaufspreis", "davon USt",
        "VK netto", "Rohertrag", "Marge", "Anmerkung"]
for i, k in enumerate(KOPF):
    z = pk.cell(row=5, column=2 + i, value=k)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    z.border = RAHMEN
pk.row_dimensions[5].height = 30

BEISPIELE = [
    ("Schokoriegel 50 g", 0.42, 0.07, 1.20, "Beispielzeile — überschreiben"),
    ("Chips 40 g", 0.38, 0.07, 1.20, ""),
    ("Müsliriegel", 0.31, 0.07, 1.00, ""),
    ("Softdrink 0,5 l PET", 0.55, 0.19, 1.50, "Pfand siehe Hinweis unten"),
    ("Wasser 0,5 l", 0.28, 0.19, 1.00, ""),
    ("Energydrink 0,25 l", 0.62, 0.19, 1.80, ""),
    ("Kaugummi", 0.35, 0.07, 1.00, ""),
]

ERSTE, LETZTE = 6, 35
for i in range(ERSTE, LETZTE + 1):
    idx = i - ERSTE
    vorgabe = BEISPIELE[idx] if idx < len(BEISPIELE) else None

    # Artikel
    z = pk.cell(row=i, column=2, value=vorgabe[0] if vorgabe else None)
    z.font = Font(name=SCHRIFT, size=10, color="0000FF")
    z.fill = PatternFill("solid", fgColor=GELB)
    z.border = RAHMEN

    # Einkauf netto, USt-Satz, Verkaufspreis brutto
    for spalte, wert, fmt in ((3, vorgabe[1] if vorgabe else None, EURO),
                              (4, vorgabe[2] if vorgabe else None, PROZENT),
                              (5, vorgabe[3] if vorgabe else None, EURO)):
        z = pk.cell(row=i, column=spalte, value=wert)
        z.number_format = fmt
        z.font = Font(name=SCHRIFT, size=10, color="0000FF")
        z.fill = PatternFill("solid", fgColor=GELB)
        z.border = RAHMEN

    # Enthaltene Umsatzsteuer aus dem Bruttopreis herausrechnen.
    z = pk.cell(row=i, column=6, value=f'=IF(E{i}="","",E{i}-E{i}/(1+D{i}))')
    z.number_format = EURO
    z.font = Font(name=SCHRIFT, size=10)
    z.border = RAHMEN

    # Verkaufspreis netto
    z = pk.cell(row=i, column=7, value=f'=IF(E{i}="","",E{i}-F{i})')
    z.number_format = EURO
    z.font = Font(name=SCHRIFT, size=10)
    z.border = RAHMEN

    # Rohertrag = netto Verkauf minus netto Einkauf. Beides netto, sonst
    # vergleicht man Preise mit und ohne Steuer.
    z = pk.cell(row=i, column=8, value=f'=IF(E{i}="","",G{i}-C{i})')
    z.number_format = EURO
    z.font = Font(name=SCHRIFT, size=10, bold=True)
    z.fill = PatternFill("solid", fgColor=GRUEN)
    z.border = RAHMEN

    # Marge auf den Nettoverkaufspreis
    z = pk.cell(row=i, column=9, value=f'=IF(OR(E{i}="",G{i}=0),"",H{i}/G{i})')
    z.number_format = PROZENT
    z.font = Font(name=SCHRIFT, size=10)
    z.border = RAHMEN

    z = pk.cell(row=i, column=10, value=vorgabe[4] if vorgabe else None)
    z.font = Font(name=SCHRIFT, size=9, color="595959")
    z.border = RAHMEN

# Auswertung
r = LETZTE + 2
z = pk.cell(row=r, column=2, value="Durchschnittlicher Rohertrag je Verkauf")
z.font = Font(name=SCHRIFT, size=12, bold=True)
z = pk.cell(row=r, column=8, value=f"=IFERROR(AVERAGE(H{ERSTE}:H{LETZTE}),0)")
z.number_format = EURO
z.font = Font(name=SCHRIFT, size=12, bold=True, color="006100")
z.fill = PatternFill("solid", fgColor=GRUEN)
z.border = RAHMEN
pk.cell(row=r, column=10,
        value="Diese Zahl übernimmt der Standort-Rechner automatisch.").font = Font(
            name=SCHRIFT, size=9, color="595959")

r += 1
z = pk.cell(row=r, column=2, value="Durchschnittlicher Verkaufspreis (brutto)")
z.font = Font(name=SCHRIFT, size=11, bold=True)
z = pk.cell(row=r, column=8, value=f"=IFERROR(AVERAGE(E{ERSTE}:E{LETZTE}),0)")
z.number_format = EURO
z.font = Font(name=SCHRIFT, size=11, bold=True)
z.border = RAHMEN

r += 1
z = pk.cell(row=r, column=2, value="Durchschnittliche Marge")
z.font = Font(name=SCHRIFT, size=11, bold=True)
z = pk.cell(row=r, column=8, value=f"=IFERROR(AVERAGE(I{ERSTE}:I{LETZTE}),0)")
z.number_format = PROZENT
z.font = Font(name=SCHRIFT, size=11, bold=True)
z.border = RAHMEN

r += 2
r = text(pk, r, "ZUM PFAND", 11, True, DUNKEL)
r = text(pk, r,
         "Pfand ist kein Ertrag. Es ist ein durchlaufender Posten: Du zahlst es beim Einkauf "
         "und bekommst es bei der Rückgabe zurück. Trag den Verkaufspreis deshalb OHNE Pfand "
         "ein und schlage das Pfand am Automaten separat auf. Wer Pfand in den Verkaufspreis "
         "rechnet, hält seine Marge für größer, als sie ist.")
r += 1
r = text(pk, r, "ZUR PREISGESTALTUNG", 11, True, DUNKEL)
r = text(pk, r,
         "Runde Preise verkaufen an Automaten besser als krumme — nicht wegen der Psychologie, "
         "sondern weil das Münzwechseln entfällt. 1,00 €, 1,50 € und 2,00 € sind deshalb "
         "üblich. Ein Preis von 1,30 € kostet dich Verkäufe von Leuten, die kein Kleingeld "
         "haben, und Wechselgeld, das du vorhalten musst.")


# ===========================================================================
# 3 — Standort-Rechner
# ===========================================================================

sr = wb.create_sheet("Standort-Rechner")
sr.sheet_view.showGridLines = False
breiten(sr, {"A": 2, "B": 46, "C": 16, "D": 58})

titel(sr, "B2", "Standort-Rechner", 16)
sr["B3"] = "Ein Standort, vollständig gerechnet. Gelbe Felder ausfüllen."
sr["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

r = 5
block(sr, r, "DER STANDORT"); r += 1
r = zeile(sr, r, "Bezeichnung", "z. B. Firma Müller, Halle 2", eingabe=True)
r = zeile(sr, r, "Art des Standorts", "Betrieb / Schule / Waschsalon / Praxis …", eingabe=True)
r = zeile(sr, r, "Personen vor Ort (täglich)", 120, ZAHL0, eingabe=True,
          hinweis="Wie viele Menschen kommen am Automaten vorbei? Nicht die "
                  "Gesamtbelegschaft — nur die, die tatsächlich vorbeilaufen.")
# Zeilennummern immer direkt nach dem Schreiben merken. Sie aus einer anderen
# Zeile hochzurechnen geht schief, sobald eine Leerzeile oder Überschrift
# dazwischenkommt — und der Fehler fällt nicht auf, weil die Formel gültig
# bleibt und nur auf eine leere Zelle zeigt.
ZEILE_PERSONEN = r - 1
r = zeile(sr, r, "Öffnungstage pro Monat", 21, ZAHL0, eingabe=True,
          hinweis="Betrieb ohne Wochenende: 21. Öffentlich zugänglich rund um die Uhr: 30.")
ZEILE_TAGE = r - 1

r += 1
block(sr, r, "ERWARTETER UMSATZ"); r += 1
r = zeile(sr, r, "Verkäufe je Person und Tag", 0.10, '0.000', eingabe=True,
          hinweis="Erfahrungswert: 0,05 bis 0,15. Bei 80 Personen und 0,08 sind das "
                  "gut 6 Verkäufe am Tag. Lieber zu vorsichtig schätzen.")
ZEILE_QUOTE = r - 1
r = zeile(sr, r, "Verkäufe pro Tag",
          f"=ROUND(C{ZEILE_PERSONEN}*C{ZEILE_QUOTE},1)", ZAHL1,
          hinweis="Personen × Quote")
ZEILE_VK_TAG = r - 1
r = zeile(sr, r, "Verkäufe pro Monat", f"=ROUND(C{ZEILE_VK_TAG}*C{ZEILE_TAGE},0)", ZAHL0)
ZEILE_VK_MON = r - 1
r = zeile(sr, r, "Rohertrag je Verkauf",
          "=Produktkalkulation!H37", EURO,
          hinweis="Kommt aus dem Blatt Produktkalkulation.")
ZEILE_ROHERTRAG = r - 1
r = zeile(sr, r, "Rohertrag pro Monat", f"=C{ZEILE_VK_MON}*C{ZEILE_ROHERTRAG}", EURO,
          fett=True, hervorheben=HELL)
ZEILE_ROH_MONAT = r - 1

r += 1
block(sr, r, "KOSTEN DES STANDORTS"); r += 1

r = zeile(sr, r, "Modell der Standortvergütung", "fest", eingabe=True,
          hinweis='Trag "fest" oder "prozentual" ein. Danach nur das passende Feld ausfüllen.')
ZEILE_MODELL = r - 1
r = zeile(sr, r, "  feste Miete pro Monat", 30.0, EURO, eingabe=True,
          hinweis="Übliche Spanne 0 bis 80 €. Viele Standorte geben die Fläche umsonst her, "
                  "wenn der Automat gewünscht ist.")
ZEILE_MIETE_FEST = r - 1
r = zeile(sr, r, "  Beteiligung am Bruttoumsatz", 0.10, PROZENT, eingabe=True,
          hinweis="Übliche Spanne 10 bis 20 %. Achtung: vom BRUTTO-Umsatz, nicht vom Gewinn.")
ZEILE_BETEILIGUNG = r - 1
r = zeile(sr, r, "  durchschnittlicher Verkaufspreis brutto",
          "=Produktkalkulation!H38", EURO,
          hinweis="Grundlage für die prozentuale Beteiligung.")
ZEILE_VK_PREIS = r - 1
r = zeile(sr, r, "Standortvergütung pro Monat",
          f'=IF(C{ZEILE_MODELL}="prozentual",'
          f'C{ZEILE_VK_MON}*C{ZEILE_VK_PREIS}*C{ZEILE_BETEILIGUNG},C{ZEILE_MIETE_FEST})',
          EURO, fett=True)
ZEILE_VERGUETUNG = r - 1

r += 1
r = zeile(sr, r, "Stromverbrauch pro Monat (kWh)", 45, ZAHL0, eingabe=True,
          hinweis="Snackautomat ohne Kühlung 15–25 kWh, mit Kühlung 40–90 kWh. "
                  "Bei Getränkekühlern im Sommer deutlich mehr.")
ZEILE_KWH = r - 1
r = zeile(sr, r, "Strompreis je kWh", 0.32, EURO, eingabe=True,
          hinweis="Zahlt der Standort den Strom, trag 0 ein — und halte das schriftlich fest.")
ZEILE_STROMPREIS = r - 1
r = zeile(sr, r, "Stromkosten pro Monat", f"=C{ZEILE_KWH}*C{ZEILE_STROMPREIS}", EURO)
ZEILE_STROM = r - 1

r += 1
block(sr, r, "DEINE FAHRTEN UND DEINE ZEIT"); r += 1
r = zeile(sr, r, "Befüllfahrten pro Monat", 2, ZAHL0, eingabe=True,
          hinweis="Zu selten befüllen kostet Umsatz, zu oft kostet Zeit.")
ZEILE_FAHRTEN = r - 1
r = zeile(sr, r, "Entfernung je Fahrt (hin und zurück, km)", 12, ZAHL0, eingabe=True)
ZEILE_KM = r - 1
r = zeile(sr, r, "Kosten je Kilometer", 0.30, EURO, eingabe=True,
          hinweis="0,30 € ist die pauschale Größe, mit der viele rechnen. Wer ein teures "
                  "Fahrzeug fährt, liegt höher.")
ZEILE_KM_SATZ = r - 1
r = zeile(sr, r, "Zeitaufwand je Fahrt (Stunden)", 0.5, ZAHL1, eingabe=True,
          hinweis="Fahrt, Befüllung, Abrechnung, Kasse leeren. Ehrlich schätzen. Gehört der Automat zu einer Tour mit mehreren Standorten, rechne nur den Anteil, der auf ihn entfällt.")
ZEILE_STUNDEN = r - 1
r = zeile(sr, r, "Dein Stundensatz", 25.0, EURO, eingabe=True,
          hinweis="Was deine Zeit wert ist. Setzt du hier 0 ein, arbeitest du umsonst — "
                  "dann rechnet die Datei den Standort schön.")
ZEILE_STUNDENSATZ = r - 1
r = zeile(sr, r, "Fahrtkosten pro Monat",
          f"=C{ZEILE_FAHRTEN}*C{ZEILE_KM}*C{ZEILE_KM_SATZ}", EURO)
ZEILE_FAHRTKOSTEN = r - 1
r = zeile(sr, r, "Wert deiner Arbeitszeit pro Monat",
          f"=C{ZEILE_FAHRTEN}*C{ZEILE_STUNDEN}*C{ZEILE_STUNDENSATZ}", EURO)
ZEILE_ZEITWERT = r - 1

r += 1
block(sr, r, "WEITERE LAUFENDE KOSTEN"); r += 1
r = zeile(sr, r, "Schwund und Verderb (% vom Rohertrag)", 0.02, PROZENT, eingabe=True,
          hinweis="Abgelaufene Ware, Diebstahl, verklemmte Artikel. 1–4 % sind normal.")
ZEILE_SCHWUND_SATZ = r - 1
r = zeile(sr, r, "Schwund pro Monat",
          f"=C{ZEILE_ROH_MONAT}*C{ZEILE_SCHWUND_SATZ}", EURO,
          hinweis="Bewusst vereinfacht auf den Rohertrag bezogen, damit nur eine Zahl "
                  "einzutragen ist.")
ZEILE_SCHWUND = r - 1
r = zeile(sr, r, "Rücklage Reparatur und Wartung", 15.0, EURO, eingabe=True,
          hinweis="Nicht weglassen. Ein Münzprüfer kostet 150–400 €, ein Kühlaggregat mehr. "
                  "15–25 € im Monat je Automat sind eine vernünftige Rücklage.")
ZEILE_REPARATUR = r - 1
r = zeile(sr, r, "Versicherung je Automat pro Monat", 4.0, EURO, eingabe=True,
          hinweis="Anteilig aus deiner Betriebs- oder Inhaltsversicherung.")
ZEILE_VERSICHERUNG = r - 1
r = zeile(sr, r, "Telemetrie / Mobilfunk pro Monat", 3.0, EURO, eingabe=True,
          hinweis="Nur wenn der Automat Füllstände meldet. Sonst 0.")
ZEILE_TELEMETRIE = r - 1
r = zeile(sr, r, "Sonstiges pro Monat", 0.0, EURO, eingabe=True)
ZEILE_SONSTIGES = r - 1

r += 1
block(sr, r, "ERGEBNIS"); r += 1
r = zeile(sr, r, "Kosten pro Monat gesamt",
          f"=C{ZEILE_VERGUETUNG}+C{ZEILE_STROM}+C{ZEILE_FAHRTKOSTEN}+C{ZEILE_ZEITWERT}"
          f"+C{ZEILE_SCHWUND}+C{ZEILE_REPARATUR}+C{ZEILE_VERSICHERUNG}"
          f"+C{ZEILE_TELEMETRIE}+C{ZEILE_SONSTIGES}",
          EURO, fett=True)
ZEILE_KOSTEN = r - 1
r = zeile(sr, r, "Gewinn pro Monat",
          f"=C{ZEILE_ROH_MONAT}-C{ZEILE_KOSTEN}", EURO, fett=True, hervorheben=GRUEN,
          hinweis="Rohertrag minus alle Kosten, deine Arbeitszeit eingerechnet.")
ZEILE_GEWINN = r - 1
r = zeile(sr, r, "Gewinn pro Jahr", f"=C{ZEILE_GEWINN}*12", EURO, fett=True)
ZEILE_GEWINN_JAHR = r - 1

r += 1
r = zeile(sr, r, "Break-even: nötige Verkäufe pro Tag",
          f"=IFERROR(ROUND(C{ZEILE_KOSTEN}/C{ZEILE_ROHERTRAG}/C{ZEILE_TAGE},1),0)",
          ZAHL1, fett=True, hervorheben=GELB,
          hinweis="Ab so vielen Verkäufen am Tag trägt sich der Standort. Liegt diese Zahl "
                  "über deiner Schätzung oben, ist der Standort nicht rentabel.")
ZEILE_BREAKEVEN = r - 1
r = zeile(sr, r, "Puffer",
          f'=IF(C{ZEILE_VK_TAG}=0,"",IFERROR(C{ZEILE_VK_TAG}/C{ZEILE_BREAKEVEN}-1,""))',
          PROZENT,
          hinweis="Wie weit du über dem Break-even liegst. Unter 20 % ist der Standort "
                  "knapp — eine Preiserhöhung beim Lieferanten kippt ihn.")
ZEILE_PUFFER = r - 1

r += 1
r = zeile(sr, r, "Anschaffungspreis des Automaten", 2500.0, EURO, eingabe=True,
          hinweis="Gebraucht 800–2.500 €, neu 3.000–8.000 €. Bei Miete oder Leasing hier 0 "
                  "eintragen und die Rate unter „Sonstiges“ erfassen.")
ZEILE_ANSCHAFFUNG = r - 1
r = zeile(sr, r, "Erstbefüllung (gebundenes Kapital)", 250.0, EURO, eingabe=True)
ZEILE_ERSTBEFUELLUNG = r - 1
r = zeile(sr, r, "Amortisation in Monaten",
          f'=IF(C{ZEILE_GEWINN}<=0,"trägt sich nicht",'
          f'ROUND((C{ZEILE_ANSCHAFFUNG}+C{ZEILE_ERSTBEFUELLUNG})/C{ZEILE_GEWINN},1))',
          ZAHL1, fett=True, hervorheben=HELL,
          hinweis="Nach so vielen Monaten hast du Automat und Erstbefüllung wieder drin.")

r += 2
r = zeile(sr, r, "Einschätzung",
          f'=IF(C{ZEILE_GEWINN}<=0,"Finger weg — dieser Standort kostet dich Geld.",'
          f'IF(C{ZEILE_GEWINN}<25,"Grenzwertig. Lohnt nur, wenn die Fahrt ohnehin ansteht.",'
          f'IF(C{ZEILE_GEWINN}<75,"Solide. Kein Selbstläufer, aber er trägt sich.",'
          f'"Guter Standort. Sichere ihn dir schriftlich.")))')
sr.cell(row=r - 1, column=3).font = Font(name=SCHRIFT, size=11, bold=True, color=DUNKEL)
sr.cell(row=r - 1, column=3).alignment = Alignment(wrap_text=True, vertical="center")
sr.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)

r += 2
r = text(sr, r,
         "Hinweis: Die Datei rechnet ohne Umsatzsteuer auf der Ergebnisseite, weil der "
         "Rohertrag bereits netto ermittelt wird. Ertragsteuern sind nicht berücksichtigt.",
         9, farbe="808080")


# ===========================================================================
# 4 — Standortvergleich
# ===========================================================================

sv = wb.create_sheet("Standortvergleich")
sv.sheet_view.showGridLines = False
breiten(sv, {"A": 2, "B": 34})
for i in range(3, 11):
    sv.column_dimensions[get_column_letter(i)].width = 15

titel(sv, "B2", "Standortvergleich", 16)
sv["B3"] = ("Bis zu acht Standorte nebeneinander. Vereinfachte Rechnung für die Vorauswahl — "
            "den Favoriten rechnest du im Standort-Rechner vollständig durch.")
sv["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

FELDER = [
    ("Bezeichnung", None, True, ["Firma Müller", "Waschsalon Bahnhof", "Praxis Dr. Weber"]),
    ("Personen täglich", ZAHL0, True, [80, 120, 45]),
    ("Öffnungstage / Monat", ZAHL0, True, [21, 30, 20]),
    ("Verkäufe je Person und Tag", '0.000', True, [0.08, 0.04, 0.10]),
    ("Rohertrag je Verkauf", EURO, True, [0.55, 0.55, 0.55]),
    ("Standortkosten / Monat", EURO, True, [30, 0, 25]),
    ("Strom / Monat", EURO, True, [14, 20, 12]),
    ("Fahrt und Zeit / Monat", EURO, True, [64, 40, 80]),
    ("Sonstige Kosten / Monat", EURO, True, [22, 22, 22]),
]

zstart = 6
for i, (bez, fmt, eingabe, vorgaben) in enumerate(FELDER):
    zr = zstart + i
    b = sv.cell(row=zr, column=2, value=bez)
    b.font = Font(name=SCHRIFT, size=10, bold=True)
    b.border = RAHMEN
    for s in range(3, 11):
        idx = s - 3
        wert = vorgaben[idx] if idx < len(vorgaben) else None
        z = sv.cell(row=zr, column=s, value=wert)
        if fmt:
            z.number_format = fmt
        z.font = Font(name=SCHRIFT, size=10, color="0000FF")
        z.fill = PatternFill("solid", fgColor=GELB)
        z.border = RAHMEN

Z_PERS, Z_TAGE, Z_QUOTE = zstart + 1, zstart + 2, zstart + 3
Z_ROH, Z_STANDORT, Z_STROM = zstart + 4, zstart + 5, zstart + 6
Z_FAHRT, Z_SONST = zstart + 7, zstart + 8

ergebnisse = [
    ("Verkäufe pro Monat", ZAHL0,
     lambda s: f"=ROUND({s}{Z_PERS}*{s}{Z_QUOTE}*{s}{Z_TAGE},0)", None),
    ("Rohertrag pro Monat", EURO, None, None),
    ("Kosten pro Monat", EURO, None, None),
    ("Gewinn pro Monat", EURO, None, GRUEN),
    ("Gewinn pro Jahr", EURO, None, None),
    ("Break-even Verkäufe / Tag", ZAHL1, None, GELB),
]

zr = zstart + len(FELDER) + 1
Z_VKM = zr
Z_ROHM = zr + 1
Z_KOSTM = zr + 2
Z_GEWM = zr + 3

for i, (bez, fmt, _, farbe) in enumerate(ergebnisse):
    r_ = zr + i
    b = sv.cell(row=r_, column=2, value=bez)
    b.font = Font(name=SCHRIFT, size=10, bold=True)
    b.border = RAHMEN
    for s_i in range(3, 11):
        s = get_column_letter(s_i)
        # Jede Ergebniszeile prüft zuerst, ob die Spalte überhaupt benutzt wird.
        # Ohne diese Klammer stünde in den fünf freien Spalten überall "0 €",
        # und der Käufer sucht nach Standorten, die es gar nicht gibt.
        leer = f'{s}{zstart}=""'
        if i == 0:
            f = f'=IF({leer},"",ROUND({s}{Z_PERS}*{s}{Z_QUOTE}*{s}{Z_TAGE},0))'
        elif i == 1:
            f = f'=IF({leer},"",{s}{Z_VKM}*{s}{Z_ROH})'
        elif i == 2:
            f = f'=IF({leer},"",{s}{Z_STANDORT}+{s}{Z_STROM}+{s}{Z_FAHRT}+{s}{Z_SONST})'
        elif i == 3:
            f = f'=IF({leer},"",{s}{Z_ROHM}-{s}{Z_KOSTM})'
        elif i == 4:
            f = f'=IF({leer},"",{s}{Z_GEWM}*12)'
        else:
            f = (f'=IF(OR({leer},{s}{Z_ROH}=0,{s}{Z_TAGE}=0),"",'
                 f'ROUND({s}{Z_KOSTM}/{s}{Z_ROH}/{s}{Z_TAGE},1))')
        z = sv.cell(row=r_, column=s_i, value=f)
        if fmt:
            z.number_format = fmt
        z.font = Font(name=SCHRIFT, size=10, bold=(i == 3))
        if farbe:
            z.fill = PatternFill("solid", fgColor=farbe)
        z.border = RAHMEN

r = zr + len(ergebnisse) + 2
r = text(sv, r, "SO LIEST DU DIE TABELLE", 11, True, DUNKEL)
r = text(sv, r,
         "Der beste Standort ist nicht der mit dem höchsten Umsatz, sondern der mit dem besten "
         "Verhältnis von Gewinn zu Aufwand. Ein Standort mit 60 € Gewinn direkt um die Ecke "
         "schlägt einen mit 90 € Gewinn und 40 km Anfahrt.")
r = text(sv, r,
         "Achte besonders auf den Break-even. Liegt er nah an deiner geschätzten Verkaufszahl, "
         "ist der Standort ein Wackelkandidat: Eine Preiserhöhung beim Großhandel oder ein "
         "Monat mit Betriebsferien kippt ihn ins Minus.")


# ===========================================================================
# 5 — Nachschlagen
# ===========================================================================

nb = wb.create_sheet("Nachschlagen")
nb.sheet_view.showGridLines = False
breiten(nb, {"A": 2, "B": 40, "C": 24, "D": 52})

titel(nb, "B2", "Erfahrungswerte und Fußangeln", 16)
nb["B3"] = ("Anhaltspunkte für die Schätzung. Keine garantierten Zahlen — jeder Standort ist "
            "anders.")
nb["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

for i, k in enumerate(["Größe", "Übliche Spanne", "Worauf es ankommt"]):
    z = nb.cell(row=5, column=2 + i, value=k)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.border = RAHMEN

WERTE = [
    ("Verkäufe je Person und Tag", "0,05 – 0,15",
     "Betriebe mit Pausenzeiten am oberen Ende, Durchgangsorte am unteren. "
     "Gibt es eine Kantine oder einen Bäcker nebenan, halbiert sich der Wert."),
    ("Rohertrag je Verkauf", "0,40 – 0,80 €",
     "Bei Snacks meist höher als bei Getränken, weil der Einkauf günstiger ist "
     "und kein Pfand anfällt."),
    ("Standortmiete fest", "0 – 80 € / Monat",
     "Viele Betriebe verlangen nichts, wenn die Belegschaft den Automaten will. "
     "Immer zuerst nach der kostenlosen Variante fragen."),
    ("Umsatzbeteiligung", "10 – 20 % vom Brutto",
     "Bei hohem Umsatz teurer als eine feste Miete. Rechne beide Modelle durch, "
     "bevor du unterschreibst."),
    ("Stromverbrauch ohne Kühlung", "15 – 25 kWh / Monat",
     "Reine Spiralautomaten brauchen fast nur Licht und Steuerung."),
    ("Stromverbrauch mit Kühlung", "40 – 90 kWh / Monat",
     "Im Sommer und in warmen Räumen am oberen Ende. Ältere Geräte deutlich darüber."),
    ("Automat gebraucht", "800 – 2.500 €",
     "Auf Ersatzteilversorgung und Münzprüfer achten. Ein Gerät ohne Ersatzteile "
     "ist beim ersten Defekt Schrott."),
    ("Automat neu", "3.000 – 8.000 €",
     "Rechnet sich meist erst ab mehreren Standorten oder bei Kartenzahlung."),
    ("Rücklage Reparatur", "15 – 25 € / Monat",
     "Münzprüfer 150–400 €, Kühlaggregat 400–900 €, Steuerplatine 200–500 €."),
    ("Schwund und Verderb", "1 – 4 %",
     "Abgelaufene Ware ist der größere Posten, nicht Diebstahl. Kurze "
     "Mindesthaltbarkeit nur bei hohem Durchsatz einkaufen."),
    ("Befüllfahrten", "1 – 4 pro Monat",
     "Je seltener, desto besser für die Rendite — aber ein leerer Automat "
     "verliert Stammkunden."),
]

zr = 6
for bez, spanne, wozu in WERTE:
    for spalte, wert in ((2, bez), (3, spanne), (4, wozu)):
        z = nb.cell(row=zr, column=spalte, value=wert)
        z.font = Font(name=SCHRIFT, size=10, bold=(spalte == 2))
        z.alignment = Alignment(wrap_text=True, vertical="top")
        z.border = RAHMEN
    nb.row_dimensions[zr].height = 32
    zr += 1

zr += 2
zr = text(nb, zr, "WORAN STANDORTE SCHEITERN", 12, True, DUNKEL)
zr += 1
for punkt in [
    "Keine schriftliche Vereinbarung. Der Ansprechpartner wechselt, der Nachfolger weiß von "
    "nichts, und der Automat steht plötzlich im Weg.",
    "Strom nicht geregelt. Wer zahlt ihn? Ohne Absprache steht das irgendwann im Raum, meist "
    "verbunden mit einer Nachforderung.",
    "Kein Zugang außerhalb der Öffnungszeiten. Wenn du nur dienstags zwischen 9 und 11 Uhr "
    "befüllen kannst, bestimmt der Standort deinen Kalender.",
    "Automat am falschen Platz. Nicht sichtbar heißt nicht gekauft. Der beste Platz ist dort, "
    "wo Menschen ohnehin warten.",
    "Keine Kartenzahlung. Der Anteil der Menschen ohne Bargeld wächst. Ein Kartenmodul kostet "
    "Geld und laufende Gebühren, kann den Umsatz aber deutlich heben — rechne es im "
    "Standort-Rechner unter „Sonstiges“ durch.",
]:
    zr = text(nb, zr, "•  " + punkt)

zr += 2
zr = text(nb, zr,
          "Arbeitshilfe, keine Steuer- oder Rechtsberatung. Alle Werte sind Erfahrungswerte "
          "und müssen für den eigenen Fall geprüft werden. Stand: August 2026.",
          9, farbe="808080")


# ---------------------------------------------------------------------------

wb.save("Automaten-Rentabilitaetsrechner.xlsx")
print("Automaten-Rentabilitaetsrechner.xlsx erzeugt")
