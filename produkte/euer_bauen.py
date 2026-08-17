"""
Erzeugt die EÜR-Vorlage für Kleinunternehmer nach § 19 UStG.

Verkaufsprodukt für den Marktplatz. Zielgruppe: Solo-Selbstständige und
Kleinunternehmer in Deutschland, die ihre Einnahmen-Überschuss-Rechnung selbst
machen wollen.

Aufbau:
  - Anleitung      Was der Käufer tun muss, plus Haftungshinweis
  - Stammdaten     Name, Steuernummer, Jahr — wird oben in die Blätter gezogen
  - Einnahmen      Erfassungsliste
  - Ausgaben       Erfassungsliste mit Kategorien
  - EÜR            Auswertung, rechnet automatisch
  - Kennzahlen     Monatsübersicht und Kleinunternehmergrenze
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCHRIFT = "Arial"

# Farben
DUNKEL = "1F3864"
HELL = "D9E2F3"
GELB = "FFF2CC"
GRAU = "F2F2F2"
GRUEN = "E2EFDA"
ROT = "FCE4E4"

EURO = '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'
DATUM = "TT.MM.JJJJ"

rand = Side(style="thin", color="BFBFBF")
RAHMEN = Border(left=rand, right=rand, top=rand, bottom=rand)

# Ausgabenkategorien nach der amtlichen Anlage EÜR
KATEGORIEN = [
    "Wareneinkauf",
    "Fremdleistungen",
    "Personalkosten",
    "Miete und Raumkosten",
    "Telefon und Internet",
    "Bürobedarf",
    "Fachliteratur und Fortbildung",
    "Werbung und Marketing",
    "Reisekosten",
    "Fahrzeugkosten",
    "Versicherungen und Beiträge",
    "Porto und Versand",
    "Software und Lizenzen",
    "Bankgebühren",
    "Steuerberatung und Recht",
    "Geringwertige Wirtschaftsgüter (bis 800 €)",
    "Abschreibungen (AfA)",
    "Sonstige Betriebsausgaben",
]

ZEILEN = 300  # Erfassungszeilen je Blatt


def titel(ws, zelle, text, groesse=14):
    ws[zelle] = text
    ws[zelle].font = Font(name=SCHRIFT, size=groesse, bold=True, color=DUNKEL)


def kopfzeile(ws, zeile, spalten, breiten):
    for i, (ueberschrift, breite) in enumerate(zip(spalten, breiten), start=1):
        z = ws.cell(row=zeile, column=i, value=ueberschrift)
        z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
        z.fill = PatternFill("solid", fgColor=DUNKEL)
        z.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        z.border = RAHMEN
        ws.column_dimensions[get_column_letter(i)].width = breite
    ws.row_dimensions[zeile].height = 28


wb = Workbook()

# ---------------------------------------------------------------------------
# Blatt 1 — Anleitung
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

titel(ws, "B2", "EÜR-Vorlage für Kleinunternehmer (§ 19 UStG)", 18)

text = [
    ("", None),
    ("So gehst du vor", "kopf"),
    ("1. Trage im Blatt „Stammdaten“ deinen Namen, deine Steuernummer und das Jahr ein.", None),
    ("2. Erfasse jede Einnahme im Blatt „Einnahmen“ — eine Zeile pro Rechnung.", None),
    ("3. Erfasse jede Ausgabe im Blatt „Ausgaben“ und wähle eine Kategorie aus der Liste.", None),
    ("4. Das Blatt „EÜR“ rechnet automatisch. Du musst dort nichts eintragen.", None),
    ("5. Das Blatt „Kennzahlen“ zeigt dir monatlich, wo du stehst.", None),
    ("", None),
    ("Welche Felder du ausfüllst", "kopf"),
    ("Gelb hinterlegte Felder sind zum Ausfüllen. Alles andere rechnet sich selbst.", None),
    ("Überschreibe keine Formeln — sonst stimmen die Summen nicht mehr.", None),
    ("", None),
    ("Was Kleinunternehmer beachten müssen", "kopf"),
    ("Als Kleinunternehmer nach § 19 UStG weist du auf deinen Rechnungen keine", None),
    ("Umsatzsteuer aus. Deshalb trägst du hier überall Bruttobeträge ein —", None),
    ("es gibt keine Vorsteuer, die du abziehen könntest.", None),
    ("", None),
    ("Die Grenzen seit der Reform 2025:", None),
    ("   • 25.000 € Umsatz im Vorjahr", None),
    ("   • 100.000 € im laufenden Jahr", None),
    ("", None),
    ("Wichtig: Die 100.000 € wirken SOFORT. Wird die Grenze mitten im Jahr", None),
    ("überschritten, ist bereits der auslösende Umsatz steuerpflichtig — nicht", None),
    ("erst das Folgejahr. Das Blatt „Kennzahlen“ warnt dich rechtzeitig.", None),
    ("", None),
    ("Belege aufbewahren", "kopf"),
    ("Diese Tabelle ersetzt keine Belege. Du musst Rechnungen und Quittungen", None),
    ("zusätzlich aufbewahren — in der Regel 10 Jahre, geordnet und nachvollziehbar.", None),
    ("Trage in der Spalte „Beleg-Nr.“ ein, wo der Beleg abgelegt ist.", None),
    ("", None),
    ("Wichtiger Hinweis", "warnung"),
    ("Diese Vorlage ist eine Arbeitshilfe, KEINE Steuerberatung. Sie ersetzt weder", None),
    ("einen Steuerberater noch die Prüfung durch das Finanzamt. Für die Richtigkeit", None),
    ("deiner Angaben gegenüber dem Finanzamt bist ausschließlich du verantwortlich.", None),
    ("Im Zweifel: frag deinen Steuerberater.", None),
    ("", None),
    ("Die genannten Grenzwerte entsprechen dem Stand August 2026. Steuerrecht", None),
    ("ändert sich — prüfe die aktuellen Werte vor der Abgabe.", None),
]

r = 3
for inhalt, art in text:
    z = ws.cell(row=r, column=2, value=inhalt)
    if art == "kopf":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color=DUNKEL)
    elif art == "warnung":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color="C00000")
    else:
        z.font = Font(name=SCHRIFT, size=11)
    z.alignment = Alignment(vertical="center")
    r += 1

# ---------------------------------------------------------------------------
# Blatt 2 — Stammdaten
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Stammdaten")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 42

titel(ws, "B2", "Stammdaten", 16)
ws["B3"] = "Gelbe Felder ausfüllen. Diese Angaben erscheinen automatisch auf allen Blättern."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

felder = [
    ("Name / Firma", "Max Mustermann"),
    ("Straße und Hausnummer", "Musterstraße 12"),
    ("PLZ und Ort", "78234 Engen"),
    ("Steuernummer", "12/345/67890"),
    ("Wirtschaftsjahr", 2026),
    ("Vorjahresumsatz (€)", 0),
]

r = 5
for beschriftung, beispiel in felder:
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11, bold=True)
    b.border = RAHMEN
    w = ws.cell(row=r, column=3, value=beispiel)
    w.font = Font(name=SCHRIFT, size=11, color="0000FF")
    w.fill = PatternFill("solid", fgColor=GELB)
    w.border = RAHMEN
    if beschriftung.endswith("(€)"):
        w.number_format = EURO
    r += 1

ws["B12"] = "Die eingetragenen Werte sind Beispiele — überschreibe sie mit deinen eigenen."
ws["B12"].font = Font(name=SCHRIFT, size=10, italic=True, color="C00000")

# ---------------------------------------------------------------------------
# Blatt 3 — Einnahmen
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Einnahmen")
ws.sheet_view.showGridLines = False

titel(ws, "A1", "Einnahmen", 16)
ws["A2"] = "Eine Zeile pro Rechnung. Als Kleinunternehmer trägst du den Bruttobetrag ein."
ws["A2"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

kopfzeile(
    ws,
    4,
    ["Datum", "Beleg-Nr.", "Kunde", "Beschreibung", "Betrag (€)", "Monat"],
    [14, 14, 28, 40, 16, 10],
)

# Beispielzeile, damit das erwartete Format sichtbar ist
beispiel = ["2026-01-15", "E-2026-001", "Beispiel GmbH", "Beratungsleistung Januar", 850.00]
for i, wert in enumerate(beispiel, start=1):
    z = ws.cell(row=5, column=i, value=wert)
    z.font = Font(name=SCHRIFT, size=10, italic=True, color="808080")
    z.border = RAHMEN
ws["A5"].number_format = DATUM
ws["E5"].number_format = EURO
ws["F5"] = "=IF(A5=\"\",\"\",MONTH(A5))"
ws["F5"].font = Font(name=SCHRIFT, size=10, italic=True, color="808080")
ws["F5"].border = RAHMEN

for r in range(6, 5 + ZEILEN):
    for c in range(1, 7):
        z = ws.cell(row=r, column=c)
        z.border = RAHMEN
        z.font = Font(name=SCHRIFT, size=10)
        if c <= 5:
            z.fill = PatternFill("solid", fgColor=GELB)
    ws.cell(row=r, column=1).number_format = DATUM
    ws.cell(row=r, column=5).number_format = EURO
    f = ws.cell(row=r, column=6, value=f'=IF(A{r}="","",MONTH(A{r}))')
    f.fill = PatternFill("solid", fgColor=GRAU)

summe = 5 + ZEILEN
ws.cell(row=summe, column=4, value="Summe Einnahmen").font = Font(
    name=SCHRIFT, size=11, bold=True
)
s = ws.cell(row=summe, column=5, value=f"=SUM(E5:E{summe - 1})")
s.font = Font(name=SCHRIFT, size=11, bold=True)
s.number_format = EURO
s.fill = PatternFill("solid", fgColor=GRUEN)
s.border = RAHMEN

ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# Blatt 4 — Ausgaben
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Ausgaben")
ws.sheet_view.showGridLines = False

titel(ws, "A1", "Ausgaben", 16)
ws["A2"] = "Kategorie aus der Liste wählen — davon hängt die Auswertung im Blatt „EÜR“ ab."
ws["A2"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

kopfzeile(
    ws,
    4,
    ["Datum", "Beleg-Nr.", "Empfänger", "Kategorie", "Beschreibung", "Betrag (€)", "Monat"],
    [14, 14, 26, 34, 32, 16, 10],
)

beispiel = [
    "2026-01-08",
    "A-2026-001",
    "Musterbedarf GmbH",
    "Bürobedarf",
    "Druckerpapier und Toner",
    64.90,
]
for i, wert in enumerate(beispiel, start=1):
    z = ws.cell(row=5, column=i, value=wert)
    z.font = Font(name=SCHRIFT, size=10, italic=True, color="808080")
    z.border = RAHMEN
ws["A5"].number_format = DATUM
ws["F5"].number_format = EURO
ws["G5"] = '=IF(A5="","",MONTH(A5))'
ws["G5"].font = Font(name=SCHRIFT, size=10, italic=True, color="808080")
ws["G5"].border = RAHMEN

for r in range(6, 5 + ZEILEN):
    for c in range(1, 8):
        z = ws.cell(row=r, column=c)
        z.border = RAHMEN
        z.font = Font(name=SCHRIFT, size=10)
        if c <= 6:
            z.fill = PatternFill("solid", fgColor=GELB)
    ws.cell(row=r, column=1).number_format = DATUM
    ws.cell(row=r, column=6).number_format = EURO
    f = ws.cell(row=r, column=7, value=f'=IF(A{r}="","",MONTH(A{r}))')
    f.fill = PatternFill("solid", fgColor=GRAU)

# Auswahlliste für die Kategorien
pruefung = DataValidation(
    type="list",
    formula1='"' + ",".join(KATEGORIEN) + '"',
    allow_blank=True,
    showDropDown=False,
)
pruefung.error = "Bitte eine Kategorie aus der Liste wählen."
pruefung.errorTitle = "Unbekannte Kategorie"
ws.add_data_validation(pruefung)
pruefung.add(f"D5:D{4 + ZEILEN}")

summe_a = 5 + ZEILEN
ws.cell(row=summe_a, column=5, value="Summe Ausgaben").font = Font(
    name=SCHRIFT, size=11, bold=True
)
s = ws.cell(row=summe_a, column=6, value=f"=SUM(F5:F{summe_a - 1})")
s.font = Font(name=SCHRIFT, size=11, bold=True)
s.number_format = EURO
s.fill = PatternFill("solid", fgColor=ROT)
s.border = RAHMEN

ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# Blatt 5 — EÜR
# ---------------------------------------------------------------------------
ws = wb.create_sheet("EÜR")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 48
ws.column_dimensions["C"].width = 20

titel(ws, "B2", "Einnahmen-Überschuss-Rechnung", 16)
ws["B3"] = "=Stammdaten!C5"
ws["B3"].font = Font(name=SCHRIFT, size=11, bold=True)
ws["B4"] = '="Steuernummer: "&Stammdaten!C8&"   |   Wirtschaftsjahr: "&Stammdaten!C9'
ws["B4"].font = Font(name=SCHRIFT, size=10, color="595959")

ws["B6"] = "Betriebseinnahmen"
ws["B6"].font = Font(name=SCHRIFT, size=12, bold=True, color="FFFFFF")
ws["B6"].fill = PatternFill("solid", fgColor=DUNKEL)
ws["C6"].fill = PatternFill("solid", fgColor=DUNKEL)

ws["B7"] = "Einnahmen aus Lieferungen und Leistungen"
ws["B7"].font = Font(name=SCHRIFT, size=11)
ws["B7"].border = RAHMEN
ws["C7"] = f"=Einnahmen!E{summe}"
ws["C7"].font = Font(name=SCHRIFT, size=11)
ws["C7"].number_format = EURO
ws["C7"].border = RAHMEN

ws["B9"] = "Betriebsausgaben"
ws["B9"].font = Font(name=SCHRIFT, size=12, bold=True, color="FFFFFF")
ws["B9"].fill = PatternFill("solid", fgColor=DUNKEL)
ws["C9"].fill = PatternFill("solid", fgColor=DUNKEL)

r = 10
for kategorie in KATEGORIEN:
    b = ws.cell(row=r, column=2, value=kategorie)
    b.font = Font(name=SCHRIFT, size=11)
    b.border = RAHMEN
    w = ws.cell(
        row=r,
        column=3,
        value=f'=SUMIF(Ausgaben!$D$5:$D${4 + ZEILEN},B{r},Ausgaben!$F$5:$F${4 + ZEILEN})',
    )
    w.font = Font(name=SCHRIFT, size=11)
    w.number_format = EURO
    w.border = RAHMEN
    r += 1

ws.cell(row=r, column=2, value="Summe Betriebsausgaben").font = Font(
    name=SCHRIFT, size=11, bold=True
)
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=HELL)
sa = ws.cell(row=r, column=3, value=f"=SUM(C10:C{r - 1})")
sa.font = Font(name=SCHRIFT, size=11, bold=True)
sa.number_format = EURO
sa.fill = PatternFill("solid", fgColor=HELL)
sa.border = RAHMEN
zeile_ausgaben = r

r += 2
ws.cell(row=r, column=2, value="Gewinn / Verlust").font = Font(
    name=SCHRIFT, size=14, bold=True, color="FFFFFF"
)
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=DUNKEL)
g = ws.cell(row=r, column=3, value=f"=C7-C{zeile_ausgaben}")
g.font = Font(name=SCHRIFT, size=14, bold=True, color="FFFFFF")
g.number_format = EURO
g.fill = PatternFill("solid", fgColor=DUNKEL)
zeile_gewinn = r

r += 2
ws.cell(row=r, column=2, value="Kontrollrechnung (muss 0,00 € ergeben)").font = Font(
    name=SCHRIFT, size=10, italic=True, color="595959"
)
k = ws.cell(row=r, column=3, value=f"=C7-C{zeile_ausgaben}-C{zeile_gewinn}")
k.font = Font(name=SCHRIFT, size=10, italic=True, color="595959")
k.number_format = EURO

r += 2
ws.cell(
    row=r,
    column=2,
    value="Diese Aufstellung ist eine Arbeitshilfe, keine Steuerberatung.",
).font = Font(name=SCHRIFT, size=10, italic=True, color="C00000")

# ---------------------------------------------------------------------------
# Blatt 6 — Kennzahlen
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Kennzahlen")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18

titel(ws, "B2", "Monatsübersicht", 16)

kopf = ["Monat", "Einnahmen", "Ausgaben", "Ergebnis"]
for i, text_ in enumerate(kopf, start=2):
    z = ws.cell(row=4, column=i, value=text_)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=DUNKEL)
    z.alignment = Alignment(horizontal="center")
    z.border = RAHMEN

monate = [
    "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember",
]

for i, monat in enumerate(monate):
    r = 5 + i
    nr = i + 1
    m = ws.cell(row=r, column=2, value=monat)
    m.font = Font(name=SCHRIFT, size=10)
    m.border = RAHMEN

    e = ws.cell(
        row=r,
        column=3,
        value=f'=SUMIF(Einnahmen!$F$5:$F${4 + ZEILEN},{nr},Einnahmen!$E$5:$E${4 + ZEILEN})',
    )
    a = ws.cell(
        row=r,
        column=4,
        value=f'=SUMIF(Ausgaben!$G$5:$G${4 + ZEILEN},{nr},Ausgaben!$F$5:$F${4 + ZEILEN})',
    )
    g = ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
    for z in (e, a, g):
        z.font = Font(name=SCHRIFT, size=10)
        z.number_format = EURO
        z.border = RAHMEN

r = 17
ws.cell(row=r, column=2, value="Gesamt").font = Font(name=SCHRIFT, size=11, bold=True)
for spalte in (3, 4, 5):
    b = get_column_letter(spalte)
    z = ws.cell(row=r, column=spalte, value=f"=SUM({b}5:{b}16)")
    z.font = Font(name=SCHRIFT, size=11, bold=True)
    z.number_format = EURO
    z.fill = PatternFill("solid", fgColor=HELL)
    z.border = RAHMEN

titel(ws, "B20", "Kleinunternehmergrenze im Blick", 14)

grenzen = [
    ("Umsatz laufendes Jahr", "=C17"),
    ("Grenze Vorjahr (25.000 €)", "=Stammdaten!C10"),
    ("Grenze laufendes Jahr", 100000),
]
r = 21
for beschriftung, wert in grenzen:
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11)
    b.border = RAHMEN
    w = ws.cell(row=r, column=3, value=wert)
    w.font = Font(name=SCHRIFT, size=11)
    w.number_format = EURO
    w.border = RAHMEN
    r += 1

ws.cell(row=25, column=2, value="Status").font = Font(name=SCHRIFT, size=11, bold=True)
status = ws.cell(
    row=25,
    column=3,
    value='=IF(C21>=C23,"GRENZE ÜBERSCHRITTEN – sofort Steuerberater fragen",'
    'IF(C21>=C23*0.8,"Achtung: über 80 % der Jahresgrenze",'
    'IF(C22>25000,"Vorjahr über 25.000 € – Kleinunternehmerstatus prüfen","Im Rahmen")))',
)
status.font = Font(name=SCHRIFT, size=11, bold=True)
status.fill = PatternFill("solid", fgColor=GELB)
status.border = RAHMEN
ws.column_dimensions["C"].width = 46

ws.cell(
    row=27,
    column=2,
    value="Die 100.000-€-Grenze wirkt sofort: Wird sie unterjährig überschritten,",
).font = Font(name=SCHRIFT, size=10, color="C00000")
ws.cell(
    row=28,
    column=2,
    value="ist bereits der auslösende Umsatz steuerpflichtig. Stand: August 2026.",
).font = Font(name=SCHRIFT, size=10, color="C00000")

wb.save("EUER-Vorlage-Kleinunternehmer.xlsx")
print("gespeichert")
