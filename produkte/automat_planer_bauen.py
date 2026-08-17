"""
Erzeugt den Wartungs- und Befüllungsplaner für Automatenaufsteller.

Verkaufsprodukt für den Marktplatz. Zielgruppe: Aufsteller mit mehreren
Automaten, die den Überblick verlieren — welcher Automat ist wann dran, welcher
Artikel läuft wo, und was kommt am Ende in der Kasse an.

Kern des Produkts: In fast jedem Automaten stehen zwei bis drei Artikel, die
praktisch niemand kauft. Sie binden Kapital, laufen ab und belegen einen
Schacht, der Geld verdienen könnte. Wer seine Verbräuche nicht mitschreibt,
merkt es nicht. Die Auswertung zeigt genau das.

Zu den Formeln: bewusst nur Funktionen aus der Excel-97-Ära (SUM, SUMIFS, IF,
IFERROR, COUNTIF, TODAY). Kein EDATE, kein XLOOKUP — Fälligkeiten werden über
Tage gerechnet, damit die Datei in Excel, LibreOffice, Numbers und Google
Tabellen gleich funktioniert.
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCHRIFT = "Arial"
DUNKEL = "1F3864"
HELL = "D9E2F3"
GELB = "FFF2CC"
GRUEN = "E2EFDA"
ROT = "FCE4E4"

EURO = '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'
ZAHL0 = '#,##0;[Red]-#,##0;"–"'
ZAHL1 = '#,##0.0;[Red]-#,##0.0;"–"'
DATUM = "TT.MM.JJJJ"

rand = Side(style="thin", color="BFBFBF")
RAHMEN = Border(left=rand, right=rand, top=rand, bottom=rand)


def titel(ws, zelle, text, groesse=14, farbe=DUNKEL):
    ws[zelle] = text
    ws[zelle].font = Font(name=SCHRIFT, size=groesse, bold=True, color=farbe)


def text(ws, r, inhalt, groesse=10, fett=False, farbe="000000", spalte=2):
    z = ws.cell(row=r, column=spalte, value=inhalt)
    z.font = Font(name=SCHRIFT, size=groesse, bold=fett, color=farbe)
    z.alignment = Alignment(wrap_text=True, vertical="top")
    return r + 1


def breiten(ws, paare):
    for spalte, breite in paare.items():
        ws.column_dimensions[spalte].width = breite


def kopfzeile(ws, zeile, spalten, start=2):
    """Tabellenkopf. spalten = [(Beschriftung, Breite), …]"""
    for i, (beschriftung, breite) in enumerate(spalten):
        s = start + i
        z = ws.cell(row=zeile, column=s, value=beschriftung)
        z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
        z.fill = PatternFill("solid", fgColor=DUNKEL)
        z.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        z.border = RAHMEN
        ws.column_dimensions[get_column_letter(s)].width = breite
    ws.row_dimensions[zeile].height = 30


def eingabe(ws, r, s, wert=None, fmt=None):
    z = ws.cell(row=r, column=s, value=wert)
    z.font = Font(name=SCHRIFT, size=10, color="0000FF")
    z.fill = PatternFill("solid", fgColor=GELB)
    z.border = RAHMEN
    if fmt:
        z.number_format = fmt
    return z


def rechnet(ws, r, s, formel, fmt=None, fett=False, farbe=None):
    z = ws.cell(row=r, column=s, value=formel)
    z.font = Font(name=SCHRIFT, size=10, bold=fett)
    z.border = RAHMEN
    if fmt:
        z.number_format = fmt
    if farbe:
        z.fill = PatternFill("solid", fgColor=farbe)
    return z


wb = Workbook()


# ===========================================================================
# 1 — Anleitung
# ===========================================================================

ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
breiten(ws, {"A": 2, "B": 96})

titel(ws, "B2", "Wartungs- und Befüllungsplaner", 18)
r = 4
r = text(ws, r, "Der Überblick über deine Automaten — wer ist dran, was läuft, was bleibt liegen.",
         12, True)
r += 1
r = text(ws, r,
         "Mit einem Automaten hat man alles im Kopf. Ab dem dritten nicht mehr. Diese Datei "
         "hält fest, wann welcher Automat gewartet werden muss, welcher Artikel an welchem "
         "Standort läuft und was am Ende wirklich in der Kasse ankommt.")
r += 1

r = text(ws, r, "DIE VIER BLÄTTER", 11, True, DUNKEL)
r += 1
for zeile_ in [
    "•  <Automaten>  —  deine Geräte mit Standort, Wartungsintervall und Fälligkeit. "
    "Einmal anlegen, dann nur noch das Wartungsdatum aktualisieren.",
    "•  <Befüllprotokoll>  —  jede Befüllung eine Zeile. Datum, Automat, Artikel, Stückzahl.",
    "•  <Kassenabrechnung>  —  was du bei jeder Leerung entnommen hast, gegen den Wareneinsatz.",
    "•  <Auswertung>  —  rechnet aus, welcher Artikel wo läuft und welcher totes Kapital ist.",
]:
    r = text(ws, r, zeile_.replace("<", "„").replace(">", "“"))
r += 1

r = text(ws, r, "SO FÄNGST DU AN", 11, True, DUNKEL)
r += 1
for schritt in [
    "1.  Blatt „Automaten“ — trag deine Geräte ein. Die Automaten-Nummer vergibst du selbst, "
    "z. B. A1, A2, A3. Sie verbindet alle anderen Blätter.",
    "2.  Nach jeder Tour: im Befüllprotokoll eine Zeile je Artikel, den du nachgefüllt hast.",
    "3.  Bei jeder Leerung: eine Zeile in der Kassenabrechnung.",
    "4.  Einmal im Monat: Blatt „Auswertung“ ansehen und Ladenhüter austauschen.",
]:
    r = text(ws, r, schritt)
r += 1

r = text(ws, r, "DIE FARBEN", 11, True, DUNKEL)
r += 1
r = text(ws, r, "Gelb hinterlegt mit blauer Schrift  =  hier trägst du ein.")
r = text(ws, r, "Weiß mit schwarzer Schrift  =  rechnet die Datei aus. Nicht überschreiben, "
                "sonst ist die Formel weg.")
r = text(ws, r, "Rot hinterlegt  =  Handlungsbedarf.")
r += 1

r = text(ws, r, "WARUM DAS BEFÜLLPROTOKOLL DIE MÜHE WERT IST", 11, True, DUNKEL)
r += 1
r = text(ws, r,
         "In fast jedem Automaten stehen zwei bis drei Artikel, die kaum jemand kauft. Sie "
         "binden dein Geld, laufen irgendwann ab und belegen einen Schacht, der verdienen "
         "könnte. Ohne Aufzeichnung merkt man das nicht — man sieht ja nur, was fehlt, nicht "
         "was steht.")
r = text(ws, r,
         "Die Auswertung zeigt dir je Artikel und je Standort, wie viel Stück im Zeitraum "
         "weggegangen sind. Was unter deiner selbst gesetzten Grenze liegt, wird als "
         "Ladenhüter markiert.")
r += 1

r = text(ws, r, "ZU DEN WARTUNGSFRISTEN", 11, True, DUNKEL)
r += 1
r = text(ws, r,
         "Die Fälligkeit wird als Richtwert über Tage gerechnet (Intervall in Monaten × 30). "
         "Das ist bewusst einfach gehalten, damit die Datei in jedem Tabellenprogramm gleich "
         "funktioniert. Für die Tourenplanung reicht das — maßgeblich bleiben die Vorgaben "
         "des Herstellers und, bei Kühlgeräten, die Hygieneanforderungen.")
r += 2
r = text(ws, r,
         "Arbeitshilfe, keine Rechts- oder Steuerberatung. Stand: August 2026.",
         9, farbe="808080")


# ===========================================================================
# 2 — Automaten
# ===========================================================================

au = wb.create_sheet("Automaten")
au.sheet_view.showGridLines = False
au.column_dimensions["A"].width = 2

titel(au, "B2", "Deine Automaten", 16)
au["B3"] = "Einmal anlegen. Nach jeder Wartung nur das Datum in Spalte G aktualisieren."
au["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

kopfzeile(au, 5, [
    ("Nr.", 8), ("Standort", 30), ("Modell", 22), ("Schächte", 10),
    ("Aufgestellt am", 15), ("Wartung alle (Monate)", 13), ("Letzte Wartung", 15),
    ("Nächste Wartung (Richtwert)", 17), ("Tage bis fällig", 13), ("Status", 16),
])

AU_ERSTE, AU_LETZTE = 6, 35
BEISPIELE = [
    ("A1", "Firma Müller, Halle 2", "Sielaff SM 6", 42, date(2026, 3, 1), 6, date(2026, 3, 1)),
    ("A2", "Waschsalon Bahnhofstr.", "Vendo V21", 30, date(2026, 4, 15), 6, date(2026, 4, 15)),
    ("A3", "Autohaus Weber, Wartebereich", "Sielaff FS 3", 36, date(2026, 6, 2), 12, date(2026, 6, 2)),
]

for i in range(AU_ERSTE, AU_LETZTE + 1):
    v = BEISPIELE[i - AU_ERSTE] if i - AU_ERSTE < len(BEISPIELE) else None
    eingabe(au, i, 2, v[0] if v else None)
    eingabe(au, i, 3, v[1] if v else None)
    eingabe(au, i, 4, v[2] if v else None)
    eingabe(au, i, 5, v[3] if v else None, ZAHL0)
    eingabe(au, i, 6, v[4] if v else None, DATUM)
    eingabe(au, i, 7, v[5] if v else None, ZAHL0)
    eingabe(au, i, 8, v[6] if v else None, DATUM)

    # Fälligkeit als Richtwert über Tage — siehe Anleitung.
    rechnet(au, i, 9,
            f'=IF(OR(H{i}="",G{i}=""),"",H{i}+G{i}*30)', DATUM)
    rechnet(au, i, 10, f'=IF(I{i}="","",I{i}-TODAY())', ZAHL0)
    rechnet(au, i, 11,
            f'=IF(J{i}="","",IF(J{i}<0,"überfällig",IF(J{i}<14,"bald fällig","in Ordnung")))')

# Rot einfärben, sobald die Frist überschritten ist. Bewusst eine bedingte
# Formatierung statt fester Farbe: die Zeilen ändern sich mit dem Datum.
au.conditional_formatting.add(
    f"J{AU_ERSTE}:J{AU_LETZTE}",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=PatternFill("solid", bgColor=ROT), font=Font(bold=True, color="C00000")),
)
au.conditional_formatting.add(
    f"J{AU_ERSTE}:J{AU_LETZTE}",
    CellIsRule(operator="between", formula=["0", "14"],
               fill=PatternFill("solid", bgColor=GELB)),
)

r = AU_LETZTE + 2
r = text(au, r, "Trag die Daten als echtes Datum ein (01.03.2026), nicht als Text — die Datei rechnet damit die Fälligkeit aus. Erscheint statt des Datums eine Zahl wie 46082, ist die Zelle nur falsch formatiert; stell sie auf Datum um.",
         9, farbe="595959")


# ===========================================================================
# 3 — Befüllprotokoll
# ===========================================================================

bp = wb.create_sheet("Befüllprotokoll")
bp.sheet_view.showGridLines = False
bp.column_dimensions["A"].width = 2

titel(bp, "B2", "Befüllprotokoll", 16)
bp["B3"] = ("Nach jeder Tour eine Zeile je nachgefülltem Artikel. Das ist die Grundlage der "
            "Auswertung — ohne diese Zeilen weiß niemand, was wirklich läuft.")
bp["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

kopfzeile(bp, 5, [
    ("Datum", 14), ("Automat", 10), ("Artikel", 30), ("nachgefüllt (Stück)", 13),
    ("Restbestand vorher", 13), ("verkauft seit letztem Mal", 15),
    ("Einkaufspreis je Stück", 13), ("Wareneinsatz", 14), ("Bemerkung", 28),
])

BP_ERSTE, BP_LETZTE = 6, 305
BP_BEISPIEL = [
    (date(2026, 8, 5), "A1", "Schokoriegel 50 g", 24, 6, 0.42, "läuft gut"),
    (date(2026, 8, 5), "A1", "Chips 40 g", 18, 12, 0.38, ""),
    (date(2026, 8, 5), "A1", "Wasser 0,5 l", 20, 4, 0.28, ""),
    (date(2026, 8, 5), "A2", "Softdrink 0,5 l", 24, 2, 0.55, "war fast leer"),
]

for i in range(BP_ERSTE, BP_LETZTE + 1):
    v = BP_BEISPIEL[i - BP_ERSTE] if i - BP_ERSTE < len(BP_BEISPIEL) else None
    eingabe(bp, i, 2, v[0] if v else None, DATUM)
    eingabe(bp, i, 3, v[1] if v else None)
    eingabe(bp, i, 4, v[2] if v else None)
    eingabe(bp, i, 5, v[3] if v else None, ZAHL0)
    eingabe(bp, i, 6, v[4] if v else None, ZAHL0)
    # Verkauft = was nachgefüllt wurde (der Schacht wird wieder voll gemacht).
    rechnet(bp, i, 7, f'=IF(E{i}="","",E{i})', ZAHL0)
    eingabe(bp, i, 8, v[5] if v else None, EURO)
    rechnet(bp, i, 9, f'=IF(OR(E{i}="",H{i}=""),"",E{i}*H{i})', EURO)
    eingabe(bp, i, 10, v[6] if v else None)

r = BP_LETZTE + 2
r = text(bp, r, "Zur Spalte „verkauft seit letztem Mal“: Wer einen Schacht wieder auffüllt, "
                "füllt genau die Menge nach, die verkauft wurde. Deshalb entspricht die "
                "Nachfüllmenge dem Absatz — vorausgesetzt, du füllst immer bis voll auf.",
         9, farbe="595959")
r = text(bp, r, "Ware, die du wegen Ablaufs entnimmst, gehört NICHT in diese Spalte. Trag sie "
                "als eigene Zeile mit dem Vermerk „Verderb“ in der Bemerkung ein, sonst "
                "erscheint ein Ladenhüter als Verkaufsschlager.",
         9, farbe="C00000")


# ===========================================================================
# 4 — Kassenabrechnung
# ===========================================================================

ka = wb.create_sheet("Kassenabrechnung")
ka.sheet_view.showGridLines = False
ka.column_dimensions["A"].width = 2

titel(ka, "B2", "Kassenabrechnung", 16)
ka["B3"] = ("Eine Zeile je Leerung. Zeigt, ob das, was in der Kasse liegt, zu dem passt, was "
            "aus dem Automaten heraus ist.")
ka["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

kopfzeile(ka, 5, [
    ("Datum", 14), ("Automat", 10), ("entnommen (€)", 14),
    ("Wechselgeld aufgefüllt (€)", 14), ("Einnahme netto Kasse", 15),
    ("Wareneinsatz laut Protokoll", 16), ("Rohertrag", 14), ("Bemerkung", 30),
])

KA_ERSTE, KA_LETZTE = 6, 155
KA_BEISPIEL = [
    (date(2026, 8, 5), "A1", 168.50, 20.00, "Zählwerk 1.842"),
    (date(2026, 8, 5), "A2", 94.00, 10.00, ""),
]

for i in range(KA_ERSTE, KA_LETZTE + 1):
    v = KA_BEISPIEL[i - KA_ERSTE] if i - KA_ERSTE < len(KA_BEISPIEL) else None
    eingabe(ka, i, 2, v[0] if v else None, DATUM)
    eingabe(ka, i, 3, v[1] if v else None)
    eingabe(ka, i, 4, v[2] if v else None, EURO)
    eingabe(ka, i, 5, v[3] if v else None, EURO)
    rechnet(ka, i, 6, f'=IF(D{i}="","",D{i}-E{i})', EURO)
    # Wareneinsatz derselben Leerung: gleiches Datum, gleicher Automat.
    rechnet(ka, i, 7,
            f'=IF(D{i}="","",SUMIFS(Befüllprotokoll!$I${BP_ERSTE}:$I${BP_LETZTE},'
            f'Befüllprotokoll!$B${BP_ERSTE}:$B${BP_LETZTE},$B{i},'
            f'Befüllprotokoll!$C${BP_ERSTE}:$C${BP_LETZTE},$C{i}))', EURO)
    rechnet(ka, i, 8, f'=IF(D{i}="","",F{i}-G{i})', EURO, fett=True, farbe=GRUEN)
    eingabe(ka, i, 9, v[4] if v else None)

r = KA_LETZTE + 2
z = ka.cell(row=r, column=2, value="Summe Rohertrag")
z.font = Font(name=SCHRIFT, size=12, bold=True)
rechnet(ka, r, 8, f"=SUM(H{KA_ERSTE}:H{KA_LETZTE})", EURO, fett=True, farbe=GRUEN)
r += 1
z = ka.cell(row=r, column=2, value="Summe Einnahmen")
z.font = Font(name=SCHRIFT, size=11, bold=True)
rechnet(ka, r, 6, f"=SUM(F{KA_ERSTE}:F{KA_LETZTE})", EURO, fett=True)


# ===========================================================================
# 5 — Auswertung
# ===========================================================================

aw = wb.create_sheet("Auswertung")
aw.sheet_view.showGridLines = False
aw.column_dimensions["A"].width = 2

titel(aw, "B2", "Auswertung", 16)
aw["B3"] = ("Welcher Artikel läuft an welchem Standort — und welcher bindet nur Kapital?")
aw["B3"].font = Font(name=SCHRIFT, size=10, color="595959")

r = 5
z = aw.cell(row=r, column=2, value="Ladenhüter-Grenze (Stück im ganzen Zeitraum)")
z.font = Font(name=SCHRIFT, size=11, bold=True)
aw.column_dimensions["B"].width = 40
eingabe(aw, r, 3, 10, ZAHL0)
GRENZE = f"$C${r}"
z = aw.cell(row=r, column=4,
            value="Alles darunter wird als Ladenhüter markiert. Setz die Grenze so, wie es "
                  "zu deinem Zeitraum passt.")
z.font = Font(name=SCHRIFT, size=9, color="595959")
aw.column_dimensions["D"].width = 46

r += 2
kopfzeile(aw, r, [
    ("Artikel", 30), ("Automat (leer = alle)", 16), ("Stück verkauft", 13),
    ("Wareneinsatz", 14), ("Bewertung", 18),
])
AW_KOPF = r
AW_ERSTE, AW_LETZTE = r + 1, r + 30

AW_BEISPIEL = [
    ("Schokoriegel 50 g", ""), ("Chips 40 g", ""), ("Wasser 0,5 l", ""),
    ("Softdrink 0,5 l", ""),
]

for i in range(AW_ERSTE, AW_LETZTE + 1):
    v = AW_BEISPIEL[i - AW_ERSTE] if i - AW_ERSTE < len(AW_BEISPIEL) else None
    eingabe(aw, i, 2, v[0] if v else None)
    eingabe(aw, i, 3, v[1] if v else None)

    # Leere Automatenspalte = über alle Automaten summieren. Der Vergleich
    # "<>" trifft jeden nicht leeren Eintrag.
    rechnet(aw, i, 4,
            f'=IF(B{i}="","",SUMIFS(Befüllprotokoll!$G${BP_ERSTE}:$G${BP_LETZTE},'
            f'Befüllprotokoll!$D${BP_ERSTE}:$D${BP_LETZTE},$B{i},'
            f'Befüllprotokoll!$C${BP_ERSTE}:$C${BP_LETZTE},IF($C{i}="","<>",$C{i})))',
            ZAHL0)
    rechnet(aw, i, 5,
            f'=IF(B{i}="","",SUMIFS(Befüllprotokoll!$I${BP_ERSTE}:$I${BP_LETZTE},'
            f'Befüllprotokoll!$D${BP_ERSTE}:$D${BP_LETZTE},$B{i},'
            f'Befüllprotokoll!$C${BP_ERSTE}:$C${BP_LETZTE},IF($C{i}="","<>",$C{i})))',
            EURO)
    rechnet(aw, i, 6,
            f'=IF(B{i}="","",IF(D{i}<{GRENZE},"Ladenhüter — austauschen",'
            f'IF(D{i}<{GRENZE}*2,"läuft mäßig","läuft gut")))')

aw.conditional_formatting.add(
    f"F{AW_ERSTE}:F{AW_LETZTE}",
    CellIsRule(operator="equal", formula=['"Ladenhüter — austauschen"'],
               fill=PatternFill("solid", bgColor=ROT), font=Font(bold=True, color="C00000")),
)

r = AW_LETZTE + 2
r = text(aw, r, "SO NUTZT DU DAS", 11, True, DUNKEL)
r += 1
r = text(aw, r,
         "Lass die Spalte „Automat“ leer, um einen Artikel über alle Standorte zu sehen. Trag "
         "eine Automaten-Nummer ein, um ihn nur an diesem Standort auszuwerten. Derselbe "
         "Artikel kann an einem Standort ein Verkaufsschlager und am nächsten ein Ladenhüter "
         "sein — genau das willst du sehen.")
r += 1
r = text(aw, r,
         "Ein Ladenhüter kostet dich doppelt: Er bindet Geld, und er belegt einen Schacht, "
         "der verdienen könnte. Tausch ihn gegen etwas aus, das am selben Standort schon "
         "läuft — oder gegen eine Variante davon.")
r += 2
r = text(aw, r,
         "Arbeitshilfe, keine Rechts- oder Steuerberatung. Stand: August 2026.",
         9, farbe="808080")


wb.save("Automaten-Wartungsplaner.xlsx")
print("Automaten-Wartungsplaner.xlsx erzeugt")
