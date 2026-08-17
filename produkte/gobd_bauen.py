"""
Erzeugt das GoBD-Belegsystem: PDF-Leitfaden plus Excel-Belegregister.

Verkaufsprodukt. Zielgruppe: Selbstständige, die ihre Belege digital ablegen
wollen, ohne dass das Finanzamt sie später verwirft.

Kern: Die GoBD sind für Laien schwer greifbar. Das Produkt übersetzt sie in
eine konkrete Ordnerstruktur, Namenskonvention und ein Register.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageBreak, PageTemplate,
    Paragraph, Spacer, Table, TableStyle,
)

# ===========================================================================
# Teil 1 — PDF-Leitfaden
# ===========================================================================
DUNKEL = colors.HexColor("#1F3864")
HELL = colors.HexColor("#D9E2F3")
GRAU = colors.HexColor("#595959")
ROT = colors.HexColor("#C00000")
GELB = colors.HexColor("#FFF2CC")

TITEL = "Belege richtig ablegen"
UNTERTITEL = "Das GoBD-Ordnersystem für Selbstständige"

stile = getSampleStyleSheet()
S = {
    "titel": ParagraphStyle("t", parent=stile["Title"], fontName="Helvetica-Bold",
                            fontSize=26, leading=31, textColor=DUNKEL,
                            alignment=TA_LEFT, spaceAfter=4),
    "untertitel": ParagraphStyle("u", parent=stile["Normal"], fontName="Helvetica",
                                 fontSize=13, leading=17, textColor=GRAU, spaceAfter=18),
    "h1": ParagraphStyle("h1", parent=stile["Heading1"], fontName="Helvetica-Bold",
                         fontSize=16, leading=20, textColor=DUNKEL,
                         spaceBefore=14, spaceAfter=8),
    "h2": ParagraphStyle("h2", parent=stile["Heading2"], fontName="Helvetica-Bold",
                         fontSize=12, leading=15, textColor=DUNKEL,
                         spaceBefore=10, spaceAfter=5),
    "text": ParagraphStyle("p", parent=stile["Normal"], fontName="Helvetica",
                           fontSize=10, leading=14.5, spaceAfter=6),
    "mono": ParagraphStyle("m", parent=stile["Normal"], fontName="Courier",
                           fontSize=9.5, leading=13.5, spaceAfter=2),
    "zelle": ParagraphStyle("z", parent=stile["Normal"], fontName="Helvetica",
                            fontSize=9, leading=12.5),
    "zellef": ParagraphStyle("zf", parent=stile["Normal"], fontName="Helvetica-Bold",
                             fontSize=9, leading=12.5),
    "klein": ParagraphStyle("k", parent=stile["Normal"], fontName="Helvetica",
                            fontSize=8.5, leading=12, textColor=GRAU),
}


def kopf_fuss(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(GRAU)
    canvas.drawString(20 * mm, 12 * mm, TITEL)
    canvas.drawRightString(190 * mm, 12 * mm, f"Seite {doc.page}")
    canvas.setStrokeColor(HELL)
    canvas.setLineWidth(0.5)
    canvas.line(20 * mm, 15 * mm, 190 * mm, 15 * mm)
    canvas.restoreState()


def kasten(text, farbe=GELB, rahmen="#D6B656"):
    t = Table([[Paragraph(text, S["zelle"])]], colWidths=[170 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), farbe),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor(rahmen)),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return t


def tabelle(kopf, zeilen, breiten):
    daten = [[Paragraph(f"<b>{h}</b>", S["zellef"]) for h in kopf]]
    for z in zeilen:
        daten.append([Paragraph(str(x), S["zelle"]) for x in z])
    t = Table(daten, colWidths=breiten, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DUNKEL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
    ]))
    return t


I = []
A = I.append

A(Spacer(1, 28 * mm))
A(Paragraph(TITEL, S["titel"]))
A(Paragraph(UNTERTITEL, S["untertitel"]))
A(Paragraph(
    "Die GoBD — die Grundsätze zur ordnungsmäßigen Führung und Aufbewahrung von "
    "Büchern, Aufzeichnungen und Unterlagen in elektronischer Form — klingen "
    "kompliziert. Praktisch laufen sie auf wenige Regeln hinaus, die man einmal "
    "einrichtet und danach nur noch befolgt.",
    S["text"]))
A(Spacer(1, 5 * mm))
A(kasten(
    "<b>Warum das wichtig ist</b><br/><br/>"
    "Werden deine Aufzeichnungen bei einer Prüfung als nicht ordnungsmäßig "
    "eingestuft, darf das Finanzamt schätzen. Schätzungen fallen selten zu "
    "deinen Gunsten aus. Der Aufwand, es von Anfang an richtig zu machen, ist "
    "ein Bruchteil dessen, was eine verworfene Buchführung kostet."))
A(PageBreak())

A(Paragraph("Die fünf Grundregeln", S["h1"]))
A(tabelle(
    ["Regel", "Was das bedeutet", "Praktisch"],
    [
        ["Nachvollziehbarkeit",
         "Ein sachverständiger Dritter muss deine Buchführung in angemessener Zeit verstehen.",
         "Jeder Beleg muss zur Buchung führen und umgekehrt."],
        ["Vollständigkeit",
         "Jeder Geschäftsvorfall wird erfasst — auch Barzahlungen und Kleinbeträge.",
         "Lückenlose Nummerierung, keine ausgelassenen Belege."],
        ["Richtigkeit",
         "Die Angaben entsprechen den tatsächlichen Verhältnissen.",
         "Betrag, Datum und Geschäftspartner müssen stimmen."],
        ["Zeitgerechtheit",
         "Erfassung zeitnah, nicht erst am Jahresende.",
         "Bargeschäfte täglich, unbare innerhalb von 10 Tagen."],
        ["Unveränderbarkeit",
         "Eine einmal erfasste Buchung darf nicht spurlos änderbar sein.",
         "Original-PDF nie überschreiben. Korrekturen als neuer Beleg."],
    ],
    [36 * mm, 76 * mm, 58 * mm]))

A(Spacer(1, 6 * mm))
A(Paragraph("Die Unveränderbarkeit ist der Knackpunkt", S["h2"]))
A(Paragraph(
    "Die meisten Fehler passieren hier. Eine Excel-Datei oder ein Word-Dokument "
    "erfüllt die Unveränderbarkeit nicht — man kann sie jederzeit ändern, ohne "
    "dass es auffällt. Deshalb gilt: <b>Der Beleg ist das PDF, nicht die "
    "Tabelle.</b> Die Tabelle ist nur deine Arbeitshilfe.",
    S["text"]))
A(kasten(
    "<b>Faustregel</b><br/><br/>"
    "Wenn du eine Datei nachträglich ändern kannst, ohne dass jemand es merkt, "
    "ist sie kein GoBD-konformer Beleg. Speichere Belege als PDF, lege sie in "
    "eine feste Ordnerstruktur und ändere sie danach nie wieder.",
    farbe=colors.HexColor("#FCE4E4"), rahmen="#C00000"))
A(PageBreak())

A(Paragraph("Die Ordnerstruktur", S["h1"]))
A(Paragraph(
    "Diese Struktur erfüllt die Anforderungen und lässt sich in fünf Minuten "
    "anlegen. Jahr, dann Art, dann Monat — nicht umgekehrt.",
    S["text"]))
A(Spacer(1, 3 * mm))
for zeile_ in [
    "Buchhaltung/",
    "  2026/",
    "    01_Ausgangsrechnungen/",
    "        2026-001_2026-01-08_MusterGmbH_1190-00.pdf",
    "        2026-002_2026-01-15_BeispielAG_450-00.pdf",
    "    02_Eingangsrechnungen/",
    "      01_Januar/",
    "        E-2026-0001_2026-01-03_Bueromarkt_64-90.pdf",
    "      02_Februar/",
    "    03_Kontoauszuege/",
    "    04_Kassenbuch/",
    "    05_Vertraege/",
    "    06_Steuer/",
    "        Umsatzsteuervoranmeldungen/",
    "        Jahresabschluss/",
]:
    A(Paragraph(zeile_.replace(" ", "&nbsp;"), S["mono"]))

A(Spacer(1, 6 * mm))
A(Paragraph("Die Namenskonvention", S["h1"]))
A(Paragraph(
    "Ein guter Dateiname macht jeden Beleg ohne Öffnen auffindbar. "
    "Diese Reihenfolge hat sich bewährt:",
    S["text"]))
A(Spacer(1, 2 * mm))
A(Paragraph(
    "<b>Belegnummer_Datum_Geschäftspartner_Betrag.pdf</b>", S["text"]))
A(Spacer(1, 2 * mm))
A(tabelle(
    ["Bestandteil", "Format", "Warum"],
    [
        ["Belegnummer", "2026-001 bzw. E-2026-0001",
         "Verbindet Beleg und Buchung. Lückenlos, nie doppelt."],
        ["Datum", "JJJJ-MM-TT, z. B. 2026-01-08",
         "Nur dieses Format sortiert im Dateimanager chronologisch."],
        ["Geschäftspartner", "ohne Leerzeichen, z. B. MusterGmbH",
         "Macht die Suche möglich."],
        ["Betrag", "Punkt durch Bindestrich ersetzt: 1190-00",
         "Punkte im Dateinamen führen zu Problemen."],
    ],
    [34 * mm, 62 * mm, 74 * mm]))
A(Spacer(1, 5 * mm))
A(kasten(
    "<b>Warum das Datum rückwärts?</b><br/><br/>"
    "Schreibst du <b>08.01.2026</b>, sortiert der Computer nach dem Tag — "
    "der 8. Januar landet neben dem 8. Juli. Schreibst du <b>2026-01-08</b>, "
    "stimmt die Reihenfolge automatisch. Diese eine Gewohnheit spart dir später "
    "Stunden."))
A(PageBreak())

A(Paragraph("Aufbewahrungsfristen", S["h1"]))
A(tabelle(
    ["Unterlage", "Frist", "Ab wann"],
    [
        ["Rechnungen (ein- und ausgehend)", "10 Jahre", "Ende des Jahres der Ausstellung"],
        ["Kontoauszüge", "10 Jahre", "Ende des Kalenderjahres"],
        ["Kassenbücher und Belege", "10 Jahre", "Ende des Kalenderjahres"],
        ["Jahresabschluss, EÜR", "10 Jahre", "Ende des Erstellungsjahres"],
        ["Handels- und Geschäftsbriefe", "6 Jahre", "Ende des Kalenderjahres"],
        ["Angebote, die zu keinem Auftrag führten", "6 Jahre", "Ende des Kalenderjahres"],
    ],
    [78 * mm, 32 * mm, 60 * mm]))
A(Spacer(1, 4 * mm))
A(Paragraph(
    "Die Frist beginnt erst mit dem Ende des Kalenderjahres. Eine Rechnung vom "
    "Januar 2026 muss also bis Ende 2036 aufbewahrt werden — fast elf Jahre.",
    S["text"]))

A(Spacer(1, 6 * mm))
A(Paragraph("Papierbelege einscannen", S["h1"]))
A(Paragraph(
    "Das ersetzende Scannen ist erlaubt: Nach dem Scannen darfst du das Papier "
    "vernichten — mit Ausnahmen. Voraussetzung ist eine dokumentierte "
    "Verfahrensweise.",
    S["text"]))
A(tabelle(
    ["Schritt", "Was zu tun ist"],
    [
        ["1. Scannen", "Vollständig, lesbar, farbig wenn die Farbe Bedeutung hat (z. B. Sichtvermerke)."],
        ["2. Prüfen", "Bildliche Übereinstimmung mit dem Original kontrollieren."],
        ["3. Ablegen", "Als PDF in die Ordnerstruktur, Namenskonvention einhalten."],
        ["4. Erfassen", "Im Belegregister eintragen."],
        ["5. Vernichten", "Papier darf weg — außer bei Unterlagen mit gesetzlicher Originalpflicht (z. B. notarielle Urkunden, Zollbelege)."],
    ],
    [30 * mm, 140 * mm]))
A(Spacer(1, 4 * mm))
A(kasten(
    "<b>Verfahrensdokumentation</b><br/><br/>"
    "Formal verlangen die GoBD eine Beschreibung, wie du vorgehst. Für "
    "Kleinunternehmer reicht in der Regel eine Seite: welche Software, welche "
    "Ordnerstruktur, wer scannt, wie geprüft wird, wo gesichert wird. "
    "Die beiliegende Excel-Datei enthält dafür ein ausfüllbares Blatt."))
A(PageBreak())

A(Paragraph("Die häufigsten Fehler", S["h1"]))
A(Paragraph(
    "<b>1. Belege nur im E-Mail-Postfach lassen.</b> Eine Rechnung als "
    "Anhang in einer Mail ist nicht abgelegt. Sie muss in die Ordnerstruktur.<br/><br/>"
    "<b>2. Dateien nachträglich umbenennen oder überschreiben.</b> Das "
    "verletzt die Unveränderbarkeit. Korrekturen kommen als neuer Beleg dazu.<br/><br/>"
    "<b>3. Lücken in der Belegnummerierung.</b> Fehlt eine Nummer, muss "
    "nachvollziehbar sein, warum. Storno besser als Löschen.<br/><br/>"
    "<b>4. Kein Backup.</b> Ein defektes Laufwerk entschuldigt keine "
    "fehlenden Belege. Mindestens eine zweite Kopie, besser an einem anderen Ort.<br/><br/>"
    "<b>5. Privates und Betriebliches mischen.</b> Wer im selben Ordner "
    "private Quittungen ablegt, lädt zu Rückfragen ein.<br/><br/>"
    "<b>6. Erst am Jahresende sortieren.</b> Das verletzt die Zeitgerechtheit "
    "und ist außerdem die unangenehmste Art, ein Wochenende zu verbringen.",
    S["text"]))
A(Spacer(1, 8 * mm))
A(kasten(
    "Dieser Leitfaden ist eine Arbeitshilfe und ersetzt keine Steuerberatung. "
    "Die Anforderungen im Einzelfall hängen von Art und Umfang deiner Tätigkeit "
    "ab. Rechtsstand: August 2026."))

doc = BaseDocTemplate(
    "GoBD-Belegsystem-Leitfaden.pdf", pagesize=A4,
    leftMargin=20 * mm, rightMargin=20 * mm,
    topMargin=18 * mm, bottomMargin=20 * mm,
    title=TITEL, subject=UNTERTITEL,
)
rahmen = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="n")
doc.addPageTemplates([PageTemplate(id="s", frames=rahmen, onPage=kopf_fuss)])
doc.build(I)
print("PDF gespeichert")

# ===========================================================================
# Teil 2 — Excel-Belegregister
# ===========================================================================
SCHRIFT = "Arial"
XD, XH, XG, XGR = "1F3864", "D9E2F3", "FFF2CC", "F2F2F2"
EURO = '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'
DATUM = "TT.MM.JJJJ"
srand = Side(style="thin", color="BFBFBF")
XRAHMEN = Border(left=srand, right=srand, top=srand, bottom=srand)
REG_ZEILEN = 400

wb = Workbook()

# --- Anleitung -------------------------------------------------------------
# Bewusst auch hier, obwohl der ausführliche Leitfaden als PDF beiliegt:
# Wer die Tabelle einzeln öffnet, muss ohne das PDF zurechtkommen.
ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

ws["B2"] = "Belegregister — so arbeitest du damit"
ws["B2"].font = Font(name=SCHRIFT, size=18, bold=True, color=XD)

anleitung = [
    ("", None),
    ("Diese Datei gehört zum Leitfaden „Belege richtig ablegen“ (PDF).", None),
    ("Dort steht ausführlich, warum die Regeln gelten. Hier steht, was zu tun ist.", None),
    ("", None),
    ("In drei Schritten", "kopf"),
    ("1. Beleg als PDF in deine Ordnerstruktur legen und nach dem Muster benennen:", None),
    ("   Belegnummer_Datum_Geschäftspartner_Betrag.pdf", None),
    ("2. Im Blatt „Belegregister“ eine Zeile ausfüllen — sofort, nicht später.", None),
    ("3. Die Spalte „Prüfung“ muss „ok“ zeigen. Steht dort etwas anderes, fehlt etwas.", None),
    ("", None),
    ("Was die Prüfspalte erkennt", "kopf"),
    ("   • Belegnr. DOPPELT — die Nummer gibt es schon. Nummern müssen einmalig sein.", None),
    ("   • Pflichtfeld fehlt — Datum, Geschäftspartner oder Betrag ist leer.", None),
    ("   • Dateiname fehlt — ohne ihn findest du den Beleg später nicht wieder.", None),
    ("", None),
    ("Das Blatt „Verfahrensdoku“", "kopf"),
    ("Die GoBD verlangen eine Beschreibung, wie du vorgehst. Für Kleinunternehmer", None),
    ("reicht diese eine Seite. Trag deine eigenen Angaben ein und drucke sie einmal", None),
    ("jährlich aus — sie gehört zu deinen Unterlagen.", None),
    ("", None),
    ("Der wichtigste Punkt", "warnung"),
    ("Diese Excel-Datei ist deine Erfassungshilfe, NICHT der Beleg selbst.", None),
    ("Der Beleg ist immer das PDF. Eine Tabelle lässt sich jederzeit ändern und", None),
    ("erfüllt die geforderte Unveränderbarkeit nicht.", None),
    ("", None),
    ("Drucke das Register mindestens einmal im Monat aus oder sichere es als PDF.", None),
    ("Diese Fassung bewahrst du unverändert auf.", None),
    ("", None),
    ("Gelbe Felder sind zum Ausfüllen. Graue Felder rechnen automatisch.", None),
    ("", None),
    ("Arbeitshilfe, keine Steuerberatung. Stand: August 2026.", "warnung"),
]

r = 3
for inhalt, art in anleitung:
    z = ws.cell(row=r, column=2, value=inhalt)
    if art == "kopf":
        z.font = Font(name=SCHRIFT, size=12, bold=True, color=XD)
    elif art == "warnung":
        z.font = Font(name=SCHRIFT, size=11, bold=True, color="C00000")
    else:
        z.font = Font(name=SCHRIFT, size=11)
    r += 1

ws = wb.create_sheet("Belegregister")
ws.sheet_view.showGridLines = False
ws["A1"] = "Belegregister"
ws["A1"].font = Font(name=SCHRIFT, size=16, bold=True, color=XD)
ws["A2"] = "Jeden Beleg hier eintragen — sofort, nicht am Jahresende."
ws["A2"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

spalten = [("Belegnr.", 16), ("Datum", 13), ("Art", 20), ("Geschäftspartner", 30),
           ("Beschreibung", 34), ("Betrag (€)", 15), ("Dateiname", 42),
           ("Papier vernichtet", 16), ("Prüfung", 24)]
for i, (h, b) in enumerate(spalten, start=1):
    z = ws.cell(row=4, column=i, value=h)
    z.font = Font(name=SCHRIFT, size=10, bold=True, color="FFFFFF")
    z.fill = PatternFill("solid", fgColor=XD)
    z.alignment = Alignment(horizontal="center", wrap_text=True)
    z.border = XRAHMEN
    ws.column_dimensions[get_column_letter(i)].width = b
ws.row_dimensions[4].height = 26

bsp = ["E-2026-0001", "2026-01-03", "Eingangsrechnung", "Büromarkt GmbH",
       "Druckerpapier und Toner", 64.90,
       "E-2026-0001_2026-01-03_Bueromarkt_64-90.pdf", "ja"]
for i, w in enumerate(bsp, start=1):
    z = ws.cell(row=5, column=i, value=w)
    z.font = Font(name=SCHRIFT, size=9, italic=True, color="808080")

erste, letzte = 5, 4 + REG_ZEILEN
for r in range(erste, letzte + 1):
    for c in range(1, 10):
        z = ws.cell(row=r, column=c)
        z.border = XRAHMEN
        z.font = Font(name=SCHRIFT, size=9)
        if c <= 8:
            z.fill = PatternFill("solid", fgColor=XG)
    ws.cell(row=r, column=2).number_format = DATUM
    ws.cell(row=r, column=6).number_format = EURO
    p = ws.cell(row=r, column=9, value=(
        f'=IF(A{r}="","",'
        f'IF(COUNTIF($A${erste}:$A${letzte},A{r})>1,"Belegnr. DOPPELT",'
        f'IF(OR(B{r}="",D{r}="",F{r}=""),"Pflichtfeld fehlt",'
        f'IF(G{r}="","Dateiname fehlt","ok"))))'))
    p.fill = PatternFill("solid", fgColor=XGR)
    p.font = Font(name=SCHRIFT, size=9, bold=True)

arten = ["Eingangsrechnung", "Ausgangsrechnung", "Kontoauszug", "Kassenbeleg",
         "Vertrag", "Steuerbescheid", "Sonstiges"]
dv = DataValidation(type="list", formula1='"' + ",".join(arten) + '"',
                    allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add(f"C{erste}:C{letzte}")

dv2 = DataValidation(type="list", formula1='"ja,nein,Original behalten"',
                     allow_blank=True, showDropDown=False)
ws.add_data_validation(dv2)
dv2.add(f"H{erste}:H{letzte}")

ws.freeze_panes = "A5"

r = letzte + 2
ws.cell(row=r, column=5, value="Belege erfasst").font = Font(name=SCHRIFT, size=11, bold=True)
c1 = ws.cell(row=r, column=6, value=f'=COUNTIF(A{erste}:A{letzte},"<>")')
c1.font = Font(name=SCHRIFT, size=11, bold=True)
c1.fill = PatternFill("solid", fgColor=XH)
c1.border = XRAHMEN

ws.cell(row=r + 1, column=5, value="davon beanstandet").font = Font(name=SCHRIFT, size=11, bold=True)
c2 = ws.cell(row=r + 1, column=6,
             value=f'=COUNTIFS(I{erste}:I{letzte},"<>ok",I{erste}:I{letzte},"<>")')
c2.font = Font(name=SCHRIFT, size=11, bold=True)
c2.fill = PatternFill("solid", fgColor="FCE4E4")
c2.border = XRAHMEN

# --- Verfahrensdokumentation ---
ws = wb.create_sheet("Verfahrensdoku")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 62

ws["B2"] = "Verfahrensdokumentation"
ws["B2"].font = Font(name=SCHRIFT, size=16, bold=True, color=XD)
ws["B3"] = "Die GoBD verlangen eine Beschreibung deines Vorgehens. Diese Seite reicht für Kleinunternehmer meist aus."
ws["B3"].font = Font(name=SCHRIFT, size=10, italic=True, color="595959")

punkte = [
    ("Unternehmen", "Max Mustermann, Musterstraße 12, 78234 Engen"),
    ("Steuernummer", "12/345/67890"),
    ("Gültig ab", "2026-01-01"),
    ("Verwendete Software", "Tabellenkalkulation, PDF-Ablage im Dateisystem"),
    ("Ort der Ablage", "Buchhaltung/<Jahr>/ auf dem Arbeitsrechner"),
    ("Ordnerstruktur", "Jahr / Belegart / Monat, siehe Leitfaden"),
    ("Namenskonvention", "Belegnummer_Datum_Partner_Betrag.pdf"),
    ("Wer erfasst", "Inhaber selbst"),
    ("Wann erfasst", "Bargeschäfte täglich, unbare innerhalb von 10 Tagen"),
    ("Scannen", "Papierbelege werden vollständig als PDF gescannt und geprüft"),
    ("Papiervernichtung", "Nach Prüfung, außer bei Originalpflicht"),
    ("Datensicherung", "Wöchentlich auf externe Festplatte, monatlich Cloud"),
    ("Unveränderbarkeit", "PDF wird nach Ablage nicht mehr geändert; Korrekturen als neuer Beleg"),
    ("Aufbewahrungsdauer", "10 Jahre ab Ende des Kalenderjahres"),
]
r = 5
for beschriftung, wert in punkte:
    b = ws.cell(row=r, column=2, value=beschriftung)
    b.font = Font(name=SCHRIFT, size=11, bold=True)
    b.border = XRAHMEN
    b.alignment = Alignment(vertical="center")
    w = ws.cell(row=r, column=3, value=wert)
    w.font = Font(name=SCHRIFT, size=10, color="0000FF")
    w.fill = PatternFill("solid", fgColor=XG)
    w.border = XRAHMEN
    w.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

r += 1
ws.cell(row=r, column=2,
        value="Die eingetragenen Werte sind Beispiele — an deine Verhältnisse anpassen.").font = Font(
    name=SCHRIFT, size=10, italic=True, color="C00000")
ws.cell(row=r + 1, column=2,
        value="Arbeitshilfe, keine Steuerberatung. Stand: August 2026.").font = Font(
    name=SCHRIFT, size=9, italic=True, color="595959")

wb.save("GoBD-Belegregister.xlsx")
print("Excel gespeichert")
